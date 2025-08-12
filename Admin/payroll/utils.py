from django.utils import timezone
from django.db.models import Q, Sum, Count, Avg
from django.core.exceptions import ValidationError
from accounts.models import CustomUser, SystemConfiguration, Role
from employees.models import EmployeeProfile, Contract
from attendance.models import (
    MonthlyAttendanceSummary,
    Attendance,
    LeaveRequest,
    Holiday,
)
from attendance.utils import (
    TimeCalculator,
    EmployeeDataManager,
    MonthlyCalculator,
)
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Tuple, Optional, Any, Union
import calendar
import logging
import uuid
import json

logger = logging.getLogger(__name__)


class PayrollDataProcessor:
    @staticmethod
    def get_payroll_month_dates(year: int, month: int) -> Dict[str, date]:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])

        return {
            "month_start": month_start,
            "month_end": month_end,
            "total_days": (month_end - month_start).days + 1,
        }

    @staticmethod
    def get_working_days_in_month(year: int, month: int) -> int:
        month_dates = PayrollDataProcessor.get_payroll_month_dates(year, month)
        working_days = 0

        current_date = month_dates["month_start"]
        while current_date <= month_dates["month_end"]:
            if current_date.weekday() < 5 and not Holiday.is_holiday_date(current_date):
                working_days += 1
            current_date += timedelta(days=1)

        return working_days

    @staticmethod
    def validate_payroll_period(year: int, month: int) -> Tuple[bool, str]:
        if month < 1 or month > 12:
            return False, "Invalid month. Must be between 1 and 12"

        current_date = timezone.now().date()
        if year > current_date.year or (
            year == current_date.year and month > current_date.month
        ):
            return False, "Cannot process payroll for future periods"

        if year < 2020:
            return False, "Invalid year. Must be 2020 or later"

        return True, "Valid payroll period"

    @staticmethod
    def get_employee_monthly_summary(
        employee: CustomUser, year: int, month: int
    ) -> Optional[MonthlyAttendanceSummary]:
        try:
            return MonthlyAttendanceSummary.objects.get(
                employee=employee, year=year, month=month
            )
        except MonthlyAttendanceSummary.DoesNotExist:
            summary = MonthlyAttendanceSummary.generate_for_employee_month(
                employee, year, month
            )
            return summary

    @staticmethod
    def safe_decimal_conversion(
        value: Any, default: Decimal = Decimal("0.00")
    ) -> Decimal:
        try:
            if isinstance(value, Decimal):
                return value
            if isinstance(value, (int, float)):
                return Decimal(str(value)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            if isinstance(value, str) and value.strip():
                return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return default
        except (ValueError, TypeError, InvalidOperation):
            return default


class PayrollCalculator:
    @staticmethod
    def calculate_basic_salary_components(
        employee: CustomUser, monthly_summary: MonthlyAttendanceSummary
    ) -> Dict[str, Decimal]:
        profile = EmployeeDataManager.get_employee_profile(employee)
        contract = EmployeeDataManager.get_employee_current_contract(employee)

        if not profile or not contract:
            raise ValidationError(
                f"Missing profile or contract for employee {employee.employee_code}"
            )

        basic_salary = profile.basic_salary
        working_days = monthly_summary.working_days

        expected_hours = Decimal(
            SystemConfiguration.get_setting("NET_WORKING_HOURS", "9.75")
        )
        total_expected_hours = expected_hours * working_days
        actual_hours = TimeCalculator.duration_to_decimal_hours(
            monthly_summary.total_work_time
        )

        salary_ratio = (
            min(actual_hours / total_expected_hours, Decimal("1.0"))
            if total_expected_hours > 0
            else Decimal("0.0")
        )

        daily_salary = (basic_salary / Decimal("30")).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        return {
            "basic_salary": basic_salary,
            "prorated_salary": (basic_salary * salary_ratio).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            ),
            "daily_salary": daily_salary,
            "hourly_rate": (
                (basic_salary / (working_days * expected_hours)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                if working_days > 0
                else Decimal("0.00")
            ),
            "salary_ratio": salary_ratio.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            ),
        }

    @staticmethod
    def calculate_overtime_pay(
        employee: CustomUser,
        monthly_summary: MonthlyAttendanceSummary,
        hourly_rate: Decimal,
    ) -> Dict[str, Decimal]:
        overtime_multiplier = Decimal(
            SystemConfiguration.get_setting("OVERTIME_RATE_MULTIPLIER", "1.5")
        )
        weekend_multiplier = Decimal(
            SystemConfiguration.get_setting("WEEKEND_OVERTIME_MULTIPLIER", "2.0")
        )

        regular_overtime_hours = TimeCalculator.duration_to_decimal_hours(
            monthly_summary.total_overtime
        )

        weekend_overtime_hours = Decimal("0.00")
        if SystemConfiguration.get_bool_setting("ALLOW_WEEKEND_OVERTIME", True):
            weekend_records = Attendance.objects.filter(
                employee=employee,
                date__year=monthly_summary.year,
                date__month=monthly_summary.month,
                is_weekend=True,
                overtime__gt=timedelta(0),
            )
            weekend_overtime_hours = sum(
                TimeCalculator.duration_to_decimal_hours(record.overtime)
                for record in weekend_records
            )

        regular_overtime_pay = (
            regular_overtime_hours * hourly_rate * overtime_multiplier
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        weekend_overtime_pay = (
            weekend_overtime_hours * hourly_rate * weekend_multiplier
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        return {
            "regular_overtime_hours": regular_overtime_hours,
            "weekend_overtime_hours": weekend_overtime_hours,
            "regular_overtime_pay": regular_overtime_pay,
            "weekend_overtime_pay": weekend_overtime_pay,
            "total_overtime_pay": regular_overtime_pay + weekend_overtime_pay,
        }

    @staticmethod
    def calculate_allowances(
        employee: CustomUser, monthly_summary: MonthlyAttendanceSummary
    ) -> Dict[str, Decimal]:
        role_name = employee.role.name if employee.role else "OTHER_STAFF"

        transport_allowance = Decimal(
            SystemConfiguration.get_setting(
                f"{role_name}_TRANSPORT_ALLOWANCE",
                SystemConfiguration.get_setting(
                    "DEFAULT_TRANSPORT_ALLOWANCE", "2000.00"
                ),
            )
        )

        meal_allowance = Decimal(
            SystemConfiguration.get_setting(
                f"{role_name}_MEAL_ALLOWANCE",
                SystemConfiguration.get_setting("DEFAULT_MEAL_ALLOWANCE", "1500.00"),
            )
        )

        telephone_allowance = Decimal(
            SystemConfiguration.get_setting(
                f"{role_name}_TELEPHONE_ALLOWANCE",
                SystemConfiguration.get_setting(
                    "DEFAULT_TELEPHONE_ALLOWANCE", "500.00"
                ),
            )
        )

        fuel_allowance = Decimal(
            SystemConfiguration.get_setting(
                f"{role_name}_FUEL_ALLOWANCE",
                SystemConfiguration.get_setting("DEFAULT_FUEL_ALLOWANCE", "0.00"),
            )
        )

        attendance_bonus = PayrollCalculator.calculate_attendance_bonus(
            employee, monthly_summary
        )
        performance_bonus = PayrollCalculator.calculate_performance_bonus(
            employee, monthly_summary
        )

        return {
            "transport_allowance": transport_allowance,
            "meal_allowance": meal_allowance,
            "telephone_allowance": telephone_allowance,
            "fuel_allowance": fuel_allowance,
            "attendance_bonus": attendance_bonus,
            "performance_bonus": performance_bonus,
            "total_allowances": transport_allowance
            + meal_allowance
            + telephone_allowance
            + fuel_allowance
            + attendance_bonus
            + performance_bonus,
        }

    @staticmethod
    def calculate_attendance_bonus(
        employee: CustomUser, monthly_summary: MonthlyAttendanceSummary
    ) -> Decimal:
        bonus_threshold = Decimal(
            SystemConfiguration.get_setting("ATTENDANCE_BONUS_THRESHOLD", "95.0")
        )
        bonus_amount = Decimal(
            SystemConfiguration.get_setting("ATTENDANCE_BONUS_AMOUNT", "1000.00")
        )

        if monthly_summary.attendance_percentage >= bonus_threshold:
            return bonus_amount
        return Decimal("0.00")

    @staticmethod
    def calculate_performance_bonus(
        employee: CustomUser, monthly_summary: MonthlyAttendanceSummary
    ) -> Decimal:
        punctuality_threshold = Decimal(
            SystemConfiguration.get_setting("PUNCTUALITY_BONUS_THRESHOLD", "98.0")
        )
        bonus_amount = Decimal(
            SystemConfiguration.get_setting("PUNCTUALITY_BONUS_AMOUNT", "500.00")
        )

        if monthly_summary.punctuality_score >= punctuality_threshold:
            return bonus_amount
        return Decimal("0.00")


class PayrollValidationHelper:
    @staticmethod
    def validate_employee_for_payroll(
        employee: CustomUser, year: int, month: int
    ) -> Tuple[bool, str]:
        if not employee.is_active:
            return False, f"Employee {employee.employee_code} is not active"

        if employee.status != "ACTIVE":
            return (
                False,
                f"Employee {employee.employee_code} status is {employee.status}",
            )

        profile = EmployeeDataManager.get_employee_profile(employee)
        if not profile or not profile.is_active:
            return False, f"Employee {employee.employee_code} profile is not active"

        contract = EmployeeDataManager.get_employee_current_contract(employee)
        if not contract:
            return (
                False,
                f"No active contract found for employee {employee.employee_code}",
            )

        payroll_date = date(year, month, 1)
        if employee.hire_date and payroll_date < employee.hire_date:
            return (
                False,
                f"Payroll date is before employee {employee.employee_code} hire date",
            )

        if employee.termination_date and payroll_date > employee.termination_date:
            return (
                False,
                f"Payroll date is after employee {employee.employee_code} termination date",
            )

        return True, "Employee is valid for payroll processing"

    @staticmethod
    def validate_payroll_data_integrity(
        payroll_data: Dict[str, Any],
    ) -> Tuple[bool, List[str]]:
        errors = []

        required_fields = [
            "employee",
            "year",
            "month",
            "basic_salary",
            "gross_salary",
            "net_salary",
        ]

        for field in required_fields:
            if field not in payroll_data or payroll_data[field] is None:
                errors.append(f"Missing required field: {field}")

        if "gross_salary" in payroll_data and "net_salary" in payroll_data:
            if payroll_data["gross_salary"] < payroll_data["net_salary"]:
                errors.append("Net salary cannot be greater than gross salary")

        if "basic_salary" in payroll_data and payroll_data["basic_salary"] < 0:
            errors.append("Basic salary cannot be negative")

        return len(errors) == 0, errors

    @staticmethod
    def validate_policy_compliance(
        payroll_data: Dict[str, Any],
    ) -> Tuple[bool, List[str]]:
        errors = []

        expected_daily = payroll_data["basic_salary"] / 30
        if abs(payroll_data["daily_salary"] - expected_daily) > Decimal("0.01"):
            errors.append("Daily rate calculation doesn't match policy (salary ÷ 30)")

        if payroll_data.get("working_days", 0) > 31:
            errors.append("Working days cannot exceed 31")

        if payroll_data.get("overtime_hours", 0) < 0:
            errors.append("Overtime hours cannot be negative")

        return len(errors) == 0, errors


class PayrollDeductionCalculator:
    @staticmethod
    def calculate_absence_deductions(
        employee: CustomUser,
        monthly_summary: MonthlyAttendanceSummary,
        daily_salary: Decimal,
    ) -> Dict[str, Decimal]:
        absent_deduction = (monthly_summary.absent_days * daily_salary).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        half_day_percentage = (
            Decimal(
                SystemConfiguration.get_setting("HALF_DAY_SALARY_PERCENTAGE", "50.0")
            )
            / 100
        )
        half_day_deduction = (
            monthly_summary.half_days
            * daily_salary
            * (Decimal("1.0") - half_day_percentage)
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        return {
            "absent_days": monthly_summary.absent_days,
            "absent_deduction": absent_deduction,
            "half_days": monthly_summary.half_days,
            "half_day_deduction": half_day_deduction,
            "total_absence_deduction": absent_deduction + half_day_deduction,
        }

    @staticmethod
    def calculate_policy_based_penalties(
        employee: CustomUser,
        monthly_summary: MonthlyAttendanceSummary,
        daily_salary: Decimal,
    ) -> Dict[str, Decimal]:
        role_name = employee.role.name if employee.role else "OTHER_STAFF"

        late_records = Attendance.objects.filter(
            employee=employee,
            date__year=monthly_summary.year,
            date__month=monthly_summary.month,
            status="LATE",
        )

        total_penalty = Decimal("0.00")
        full_day_deductions = 0
        half_day_deductions = 0
        penalty_details = []

        for record in late_records:
            penalty_amount = Decimal("0.00")
            penalty_type = ""

            if role_name == "OTHER_STAFF":
                grace_period = SystemConfiguration.get_int_setting(
                    "OTHER_STAFF_GRACE_PERIOD_MINUTES", 15
                )
                half_day_threshold = SystemConfiguration.get_int_setting(
                    "HALF_DAY_THRESHOLD_MINUTES", 35
                )

                if record.late_minutes >= half_day_threshold:
                    penalty_amount = daily_salary * Decimal("0.5")
                    penalty_type = "half_day_deduction"
                    half_day_deductions += 1
                elif record.late_minutes > grace_period:
                    if PayrollDeductionCalculator._check_full_day_penalty_condition(
                        record
                    ):
                        penalty_amount = daily_salary
                        penalty_type = "full_day_deduction"
                        full_day_deductions += 1
                    else:
                        penalty_amount = Decimal(
                            SystemConfiguration.get_setting(
                                "OTHER_STAFF_LATE_PENALTY_RATE", "50.00"
                            )
                        )
                        penalty_type = "late_penalty"

            elif role_name == "OFFICE_WORKER":
                half_day_threshold = SystemConfiguration.get_int_setting(
                    "HALF_DAY_THRESHOLD_MINUTES", 35
                )
                if record.late_minutes >= half_day_threshold:
                    penalty_amount = daily_salary * Decimal("0.5")
                    penalty_type = "half_day_deduction"
                    half_day_deductions += 1
                else:
                    penalty_amount = Decimal(
                        SystemConfiguration.get_setting(
                            "OFFICE_WORKER_LATE_PENALTY_RATE", "25.00"
                        )
                    )
                    penalty_type = "late_penalty"
            else:
                penalty_per_minute = Decimal(
                    SystemConfiguration.get_setting("LATE_PENALTY_PER_MINUTE", "10.00")
                )
                penalty_amount = (penalty_per_minute * record.late_minutes).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                penalty_type = "late_penalty"

            total_penalty += penalty_amount
            penalty_details.append(
                {
                    "date": record.date,
                    "late_minutes": record.late_minutes,
                    "penalty_type": penalty_type,
                    "penalty_amount": penalty_amount,
                }
            )

        return {
            "late_days": monthly_summary.late_days,
            "full_day_deductions": full_day_deductions,
            "half_day_deductions": half_day_deductions,
            "total_late_penalty": total_penalty,
            "penalty_details": penalty_details,
        }

    @staticmethod
    def _check_full_day_penalty_condition(attendance_record: Attendance) -> bool:
        if not attendance_record.first_in_time or not attendance_record.last_out_time:
            return False

        reporting_time = time(8, 0)
        grace_end_time = time(8, 15)
        penalty_check_time = time(8, 30)

        if (
            attendance_record.first_in_time >= grace_end_time
            and attendance_record.last_out_time <= penalty_check_time
        ):
            return True

        return False

    @staticmethod
    def integrate_attendance_penalties(
        employee: CustomUser, year: int, month: int, daily_salary: Decimal
    ) -> Dict[str, Decimal]:
        attendance_records = Attendance.objects.filter(
            employee=employee, date__year=year, date__month=month
        )

        total_penalties = Decimal("0.00")
        penalty_details = []

        for record in attendance_records:
            penalties = calculate_role_based_penalties(employee, record)

            if penalties.get("full_day_deduction", False):
                penalty_amount = daily_salary
                total_penalties += penalty_amount
                penalty_details.append(
                    {
                        "date": record.date,
                        "type": "full_day_deduction",
                        "amount": penalty_amount,
                    }
                )
            elif penalties.get("half_day_deduction", False):
                penalty_amount = daily_salary * Decimal("0.5")
                total_penalties += penalty_amount
                penalty_details.append(
                    {
                        "date": record.date,
                        "type": "half_day_deduction",
                        "amount": penalty_amount,
                    }
                )

            if penalties.get("late_penalty", Decimal("0.00")) > 0:
                penalty_amount = penalties["late_penalty"]
                total_penalties += penalty_amount
                penalty_details.append(
                    {
                        "date": record.date,
                        "type": "late_penalty",
                        "amount": penalty_amount,
                    }
                )

        return {"total_penalties": total_penalties, "penalty_details": penalty_details}

    @staticmethod
    def calculate_leave_deductions(
        employee: CustomUser, year: int, month: int, daily_salary: Decimal
    ) -> Dict[str, Decimal]:
        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])

        unpaid_leave_requests = LeaveRequest.objects.filter(
            employee=employee,
            status="APPROVED",
            start_date__lte=month_end,
            end_date__gte=month_start,
            leave_type__is_paid=False,
        )

        total_unpaid_days = Decimal("0.00")
        leave_details = []

        for leave_request in unpaid_leave_requests:
            leave_start = max(leave_request.start_date, month_start)
            leave_end = min(leave_request.end_date, month_end)

            if leave_request.is_half_day:
                days_in_month = Decimal("0.5")
            else:
                days_in_month = Decimal(str((leave_end - leave_start).days + 1))

            total_unpaid_days += days_in_month
            leave_details.append(
                {
                    "leave_type": leave_request.leave_type.name,
                    "start_date": leave_start,
                    "end_date": leave_end,
                    "days": days_in_month,
                    "deduction": (days_in_month * daily_salary).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    ),
                }
            )

        total_leave_deduction = (total_unpaid_days * daily_salary).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        return {
            "unpaid_leave_days": total_unpaid_days,
            "total_leave_deduction": total_leave_deduction,
            "leave_details": leave_details,
        }

    @staticmethod
    def calculate_lunch_violation_penalties(
        employee: CustomUser, year: int, month: int, daily_salary: Decimal
    ) -> Dict[str, Decimal]:
        violation_limit = SystemConfiguration.get_int_setting(
            "LUNCH_VIOLATION_LIMIT_PER_MONTH", 3
        )
        penalty_days = SystemConfiguration.get_int_setting(
            "LUNCH_VIOLATION_PENALTY_DAYS", 1
        )
        max_lunch_minutes = SystemConfiguration.get_int_setting(
            "MAX_LUNCH_DURATION_MINUTES", 75
        )

        month_start = date(year, month, 1)
        month_end = date(year, month, calendar.monthrange(year, month)[1])

        violations = Attendance.objects.filter(
            employee=employee,
            date__range=[month_start, month_end],
            break_time__gt=timedelta(minutes=max_lunch_minutes),
        ).count()

        penalty_amount = Decimal("0.00")
        if violations >= violation_limit:
            penalty_amount = (daily_salary * penalty_days).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

        return {
            "lunch_violations": violations,
            "violation_limit": violation_limit,
            "penalty_amount": penalty_amount,
        }

class PayrollTaxCalculator:
    @staticmethod
    def calculate_epf_contributions(
        gross_salary: Decimal, epf_salary_base: Decimal
    ) -> Dict[str, Decimal]:
        employee_rate = (
            Decimal(SystemConfiguration.get_setting("EPF_EMPLOYEE_RATE", "8.0")) / 100
        )
        employer_rate = (
            Decimal(SystemConfiguration.get_setting("EPF_EMPLOYER_RATE", "12.0")) / 100
        )

        employee_contribution = (epf_salary_base * employee_rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        employer_contribution = (epf_salary_base * employer_rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        return {
            "epf_salary_base": epf_salary_base,
            "employee_epf_rate": employee_rate,
            "employer_epf_rate": employer_rate,
            "employee_epf_contribution": employee_contribution,
            "employer_epf_contribution": employer_contribution,
            "total_epf_contribution": employee_contribution + employer_contribution,
        }

    @staticmethod
    def calculate_etf_contribution(gross_salary: Decimal) -> Dict[str, Decimal]:
        etf_rate = Decimal(SystemConfiguration.get_setting("ETF_RATE", "3.0")) / 100
        etf_contribution = (gross_salary * etf_rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        return {"etf_rate": etf_rate, "etf_contribution": etf_contribution}

    @staticmethod
    def calculate_income_tax(
        annual_income: Decimal, employee: CustomUser
    ) -> Dict[str, Decimal]:
        tax_free_threshold = Decimal(
            SystemConfiguration.get_setting("TAX_FREE_THRESHOLD", "1200000.00")
        )
        basic_tax_rate = (
            Decimal(SystemConfiguration.get_setting("BASIC_TAX_RATE", "6.0")) / 100
        )

        profile = EmployeeDataManager.get_employee_profile(employee)

        additional_relief = Decimal("0.00")
        if profile and profile.marital_status == "MARRIED":
            additional_relief += Decimal(
                SystemConfiguration.get_setting("SPOUSE_RELIEF", "100000.00")
            )

        if profile and profile.number_of_children > 0:
            child_relief_per_child = Decimal(
                SystemConfiguration.get_setting("CHILD_RELIEF_PER_CHILD", "75000.00")
            )
            max_children = int(
                SystemConfiguration.get_setting("MAX_CHILDREN_FOR_RELIEF", "3")
            )
            additional_relief += child_relief_per_child * min(
                profile.number_of_children, max_children
            )

        taxable_income = max(
            annual_income - tax_free_threshold - additional_relief, Decimal("0.00")
        )
        annual_tax = (taxable_income * basic_tax_rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        monthly_tax = (annual_tax / 12).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        return {
            "annual_income": annual_income,
            "tax_free_threshold": tax_free_threshold,
            "additional_relief": additional_relief,
            "taxable_income": taxable_income,
            "annual_tax": annual_tax,
            "monthly_tax": monthly_tax,
        }


class PayrollAdvanceCalculator:
    @staticmethod
    def calculate_available_advance_amount(employee: CustomUser) -> Dict[str, Decimal]:
        profile = EmployeeDataManager.get_employee_profile(employee)
        if not profile:
            return {
                "available_amount": Decimal("0.00"),
                "max_percentage": Decimal("0.00"),
            }

        max_percentage = (
            Decimal(
                SystemConfiguration.get_setting("SALARY_ADVANCE_MAX_PERCENTAGE", "50.0")
            )
            / 100
        )
        max_advance_amount = (profile.basic_salary * max_percentage).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

        current_year = timezone.now().year
        current_advances = PayrollAdvanceCalculator.get_current_advances(
            employee, current_year
        )

        total_outstanding = sum(
            advance["outstanding_amount"] for advance in current_advances
        )
        available_amount = max(max_advance_amount - total_outstanding, Decimal("0.00"))

        return {
            "basic_salary": profile.basic_salary,
            "max_percentage": max_percentage,
            "max_advance_amount": max_advance_amount,
            "current_outstanding": total_outstanding,
            "available_amount": available_amount,
            "advance_count_this_year": len(current_advances),
        }

    @staticmethod
    def get_current_advances(employee: CustomUser, year: int) -> List[Dict[str, Any]]:
        from .models import SalaryAdvance

        advances = SalaryAdvance.objects.filter(
            employee=employee, created_at__year=year, status__in=["APPROVED", "ACTIVE"]
        )

        advance_list = []
        for advance in advances:
            advance_list.append(
                {
                    "id": advance.id,
                    "amount": advance.amount,
                    "outstanding_amount": advance.outstanding_amount,
                    "monthly_deduction": advance.monthly_deduction,
                    "status": advance.status,
                }
            )

        return advance_list

    @staticmethod
    def calculate_advance_deduction(
        employee: CustomUser, year: int, month: int
    ) -> Dict[str, Decimal]:
        from .models import SalaryAdvance

        active_advances = SalaryAdvance.objects.filter(
            employee=employee, status="ACTIVE", outstanding_amount__gt=0
        )

        total_deduction = Decimal("0.00")
        advance_details = []

        for advance in active_advances:
            deduction_amount = min(
                advance.monthly_deduction, advance.outstanding_amount
            )
            total_deduction += deduction_amount

            advance_details.append(
                {
                    "advance_id": advance.id,
                    "original_amount": advance.amount,
                    "outstanding_amount": advance.outstanding_amount,
                    "monthly_deduction": advance.monthly_deduction,
                    "this_month_deduction": deduction_amount,
                }
            )

        return {
            "total_advance_deduction": total_deduction,
            "advance_details": advance_details,
            "active_advances_count": len(advance_details),
        }


class PayrollReportDataProcessor:
    @staticmethod
    def prepare_individual_payslip_data(
        employee: CustomUser, payroll_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        profile = EmployeeDataManager.get_employee_profile(employee)

        return {
            "sr_no": payroll_data.get("sr_no", ""),
            "employee_name": employee.get_full_name(),
            "division": employee.department.name if employee.department else "",
            "job_title": employee.job_title or "",
            "account_no": profile.bank_account_number if profile else "",
            "ot_basic": payroll_data.get("basic_salary", Decimal("0.00")),
            "working_days": payroll_data.get("working_days", 0),
            "basic_salary": payroll_data.get("basic_salary", Decimal("0.00")),
            "bonus_1": Decimal(SystemConfiguration.get_setting("DEFAULT_BONUS_1", "1500.00")),
            "bonus_2": Decimal(SystemConfiguration.get_setting("DEFAULT_BONUS_2", "1000.00")),
            "epf_salary": payroll_data.get("epf_salary_base", Decimal("0.00")),
            "transport_allowance": payroll_data.get("transport_allowance", Decimal("0.00")),
            "telephone_allowance": payroll_data.get("telephone_allowance", Decimal("0.00")),
            "attendance_bonus": payroll_data.get("attendance_bonus", Decimal("0.00")),
            "incentive_allowance": payroll_data.get("performance_bonus", Decimal("0.00")),
            "interim_allowance": payroll_data.get("interim_allowance", Decimal("0.00")),
            "monthly_salary": payroll_data.get("monthly_salary_total", Decimal("0.00")),
            "fuel_allowance": payroll_data.get("fuel_allowance", Decimal("0.00")),
            "meal_allowance": payroll_data.get("meal_allowance", Decimal("0.00")),
            "education_allowance": payroll_data.get("education_allowance", Decimal("0.00")),
            "religious_pay": payroll_data.get("religious_pay", Decimal("0.00")),
            "friday_salary": payroll_data.get("friday_salary", Decimal("0.00")),
            "friday_overtime": payroll_data.get("friday_overtime", Decimal("0.00")),
            "regular_overtime": payroll_data.get("regular_overtime_pay", Decimal("0.00")),
            "gross_salary": payroll_data.get("gross_salary", Decimal("0.00")),
            "working_day_meals": payroll_data.get("working_day_meals", 0),
            "overtime_hours": payroll_data.get("total_overtime_hours", Decimal("0.00")),
            "leave_days": payroll_data.get("leave_days", 0),
            "leave_deduction": payroll_data.get("leave_deduction", Decimal("0.00")),
            "late_penalty": payroll_data.get("late_penalty", Decimal("0.00")),
            "epf_deduction": payroll_data.get("epf_deduction", Decimal("0.00")),
            "total_deductions": payroll_data.get("total_deductions", Decimal("0.00")),
            "net_salary": payroll_data.get("net_salary", Decimal("0.00")),
            "fuel_per_day": Decimal(SystemConfiguration.get_setting("FUEL_PER_DAY", "50.00")),
            "meal_per_day": Decimal(SystemConfiguration.get_setting("MEAL_PER_DAY", "350.00")),
        }

    @staticmethod
    def prepare_department_summary_data(
        department_name: str, employee_payrolls: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        if not employee_payrolls:
            return {}

        summary = {
            "department_name": department_name,
            "employee_count": len(employee_payrolls),
            "total_basic_salary": sum(
                p.get("basic_salary", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_allowances": sum(
                p.get("total_allowances", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_overtime_pay": sum(
                p.get("total_overtime_pay", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_gross_salary": sum(
                p.get("gross_salary", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_deductions": sum(
                p.get("total_deductions", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_net_salary": sum(
                p.get("net_salary", Decimal("0.00")) for p in employee_payrolls
            ),
            "total_epf_contribution": sum(
                p.get("total_epf_contribution", Decimal("0.00"))
                for p in employee_payrolls
            ),
        }

        return summary

class PayrollExcelProcessor:
    @staticmethod
    def create_payroll_excel(payroll_data: List[Dict[str, Any]], year: int, month: int) -> bytes:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Payroll_{year}_{month:02d}"
        
        headers = PayrollExcelProcessor.get_dynamic_excel_headers()
        PayrollExcelProcessor._setup_excel_headers(ws, headers)
        
        departments = PayrollExcelProcessor._group_employees_by_department(payroll_data)
        current_row = PayrollExcelProcessor._write_employee_data(ws, departments, headers)
        PayrollExcelProcessor._write_totals_section(ws, departments, current_row, headers)
        PayrollExcelProcessor._format_excel_columns(ws, headers)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def get_dynamic_excel_headers():
        base_headers = [
            'Sr.No', 'Name', 'Division', 'Job Title', 'Account No', 
            'Working Days', 'Basic Salary'
        ]
        
        active_roles = Role.objects.filter(is_active=True)
        role_allowances = []
        
        for role in active_roles:
            role_name = role.name
            if SystemConfiguration.get_setting(f"{role_name}_TRANSPORT_ALLOWANCE"):
                role_allowances.append(f"{role_name} Transport")
            if SystemConfiguration.get_setting(f"{role_name}_MEAL_ALLOWANCE"):
                role_allowances.append(f"{role_name} Meal")
            if SystemConfiguration.get_setting(f"{role_name}_FUEL_ALLOWANCE"):
                role_allowances.append(f"{role_name} Fuel")
        
        standard_allowances = [
            'Transport Allowance', 'Meal Allowance', 'Telephone Allowance',
            'Fuel Allowance', 'Attendance Bonus', 'Performance Bonus'
        ]
        
        overtime_columns = [
            'Regular Overtime', 'Weekend Overtime', 'Total Overtime Hours'
        ]
        
        deduction_columns = [
            'Absent Days', 'Absent Deduction', 'Half Days', 'Half Day Deduction',
            'Late Penalties', 'Lunch Violations', 'Leave Deduction', 
            'Advance Deduction', 'EPF Deduction', 'Income Tax'
        ]
        
        summary_columns = [
            'Gross Salary', 'Total Deductions', 'Net Salary'
        ]
        
        all_headers = (base_headers + role_allowances + standard_allowances + 
                      overtime_columns + deduction_columns + summary_columns)
        
        return all_headers
    
    @staticmethod
    def _setup_excel_headers(ws, headers: List[str]):
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        header_font = Font(bold=True, size=10, name='Arial')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = border
    
    @staticmethod
    def _group_employees_by_department(payroll_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        departments = {}
        for data in payroll_data:
            dept = data.get('division', 'Unknown')
            if dept not in departments:
                departments[dept] = []
            departments[dept].append(data)
        return departments
    
    @staticmethod
    def _write_employee_data(ws, departments: Dict[str, List[Dict[str, Any]]], headers: List[str]) -> int:
        from openpyxl.styles import Border, Side, Alignment
        
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_row = 2
        for dept_name, dept_employees in departments.items():
            for emp_data in dept_employees:
                PayrollExcelProcessor._write_employee_row(ws, current_row, emp_data, border, headers)
                current_row += 1
            current_row += 1
        
        return current_row
    
    @staticmethod
    def _write_employee_row(ws, row: int, emp_data: Dict[str, Any], border, headers: List[str]):
        from openpyxl.styles import Alignment
        
        header_to_data_mapping = {
            'Sr.No': emp_data.get('sr_no', ''),
            'Name': emp_data.get('employee_name', ''),
            'Division': emp_data.get('division', ''),
            'Job Title': emp_data.get('job_title', ''),
            'Account No': emp_data.get('account_no', ''),
            'Working Days': emp_data.get('working_days', 0),
            'Basic Salary': float(emp_data.get('basic_salary', 0)),
            'Transport Allowance': float(emp_data.get('transport_allowance', 0)),
            'Meal Allowance': float(emp_data.get('meal_allowance', 0)),
            'Telephone Allowance': float(emp_data.get('telephone_allowance', 0)),
            'Fuel Allowance': float(emp_data.get('fuel_allowance', 0)),
            'Attendance Bonus': float(emp_data.get('attendance_bonus', 0)),
            'Performance Bonus': float(emp_data.get('performance_bonus', 0)),
            'Regular Overtime': float(emp_data.get('regular_overtime_pay', 0)),
            'Weekend Overtime': float(emp_data.get('weekend_overtime_pay', 0)),
            'Total Overtime Hours': float(emp_data.get('total_overtime_hours', 0)),
            'Absent Days': emp_data.get('absent_days', 0),
            'Absent Deduction': float(emp_data.get('absent_deduction', 0)),
            'Half Days': emp_data.get('half_days', 0),
            'Half Day Deduction': float(emp_data.get('half_day_deduction', 0)),
            'Late Penalties': float(emp_data.get('late_penalty', 0)),
            'Lunch Violations': float(emp_data.get('lunch_violation_penalty', 0)),
            'Leave Deduction': float(emp_data.get('leave_deduction', 0)),
            'Advance Deduction': float(emp_data.get('advance_deduction', 0)),
            'EPF Deduction': float(emp_data.get('epf_deduction', 0)),
            'Income Tax': float(emp_data.get('income_tax', 0)),
            'Gross Salary': float(emp_data.get('gross_salary', 0)),
            'Total Deductions': float(emp_data.get('total_deductions', 0)),
            'Net Salary': float(emp_data.get('net_salary', 0))
        }
        
        for col, header in enumerate(headers, 1):
            value = header_to_data_mapping.get(header, '')
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if isinstance(value, (int, float)) and col > 5:
                cell.number_format = '#,##0.00'
    
    @staticmethod
    def _write_totals_section(ws, departments: Dict[str, List[Dict[str, Any]]], start_row: int, headers: List[str]):
        from openpyxl.styles import Border, Side
        
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        current_row = start_row
        department_totals = {}
        
        for dept_name, dept_employees in departments.items():
            dept_total = PayrollExcelProcessor._calculate_department_total(dept_employees)
            department_totals[dept_name] = dept_total
            PayrollExcelProcessor._write_department_total_row(ws, current_row, f"{dept_name} Total", dept_total, border, headers)
            current_row += 2
        
        grand_total = PayrollExcelProcessor._calculate_grand_total(department_totals)
        PayrollExcelProcessor._write_department_total_row(ws, current_row, "Grand Total", grand_total, border, headers)
    
    @staticmethod
    def _calculate_department_total(dept_employees: List[Dict[str, Any]]) -> Dict[str, Any]:
        return {
            'employee_count': len(dept_employees),
            'total_basic_salary': sum(float(emp.get('basic_salary', 0)) for emp in dept_employees),
            'total_transport': sum(float(emp.get('transport_allowance', 0)) for emp in dept_employees),
            'total_meal': sum(float(emp.get('meal_allowance', 0)) for emp in dept_employees),
            'total_telephone': sum(float(emp.get('telephone_allowance', 0)) for emp in dept_employees),
            'total_fuel': sum(float(emp.get('fuel_allowance', 0)) for emp in dept_employees),
            'total_attendance_bonus': sum(float(emp.get('attendance_bonus', 0)) for emp in dept_employees),
            'total_performance_bonus': sum(float(emp.get('performance_bonus', 0)) for emp in dept_employees),
            'total_regular_ot': sum(float(emp.get('regular_overtime_pay', 0)) for emp in dept_employees),
            'total_weekend_ot': sum(float(emp.get('weekend_overtime_pay', 0)) for emp in dept_employees),
            'total_ot_hours': sum(float(emp.get('total_overtime_hours', 0)) for emp in dept_employees),
            'total_absent_days': sum(emp.get('absent_days', 0) for emp in dept_employees),
            'total_absent_deduction': sum(float(emp.get('absent_deduction', 0)) for emp in dept_employees),
            'total_half_days': sum(emp.get('half_days', 0) for emp in dept_employees),
            'total_half_day_deduction': sum(float(emp.get('half_day_deduction', 0)) for emp in dept_employees),
            'total_late_penalty': sum(float(emp.get('late_penalty', 0)) for emp in dept_employees),
            'total_lunch_violations': sum(float(emp.get('lunch_violation_penalty', 0)) for emp in dept_employees),
            'total_leave_deduction': sum(float(emp.get('leave_deduction', 0)) for emp in dept_employees),
            'total_advance_deduction': sum(float(emp.get('advance_deduction', 0)) for emp in dept_employees),
            'total_epf_deduction': sum(float(emp.get('epf_deduction', 0)) for emp in dept_employees),
            'total_income_tax': sum(float(emp.get('income_tax', 0)) for emp in dept_employees),
            'total_gross': sum(float(emp.get('gross_salary', 0)) for emp in dept_employees),
            'total_deductions': sum(float(emp.get('total_deductions', 0)) for emp in dept_employees),
            'total_net_salary': sum(float(emp.get('net_salary', 0)) for emp in dept_employees)
        }
    
    @staticmethod
    def _calculate_grand_total(department_totals: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        grand_total = {}
        numeric_fields = [
            'employee_count', 'total_basic_salary', 'total_transport', 'total_meal',
            'total_telephone', 'total_fuel', 'total_attendance_bonus', 'total_performance_bonus',
            'total_regular_ot', 'total_weekend_ot', 'total_ot_hours', 'total_absent_days',
            'total_absent_deduction', 'total_half_days', 'total_half_day_deduction',
            'total_late_penalty', 'total_lunch_violations', 'total_leave_deduction',
            'total_advance_deduction', 'total_epf_deduction', 'total_income_tax',
            'total_gross', 'total_deductions', 'total_net_salary'
        ]
        
        for field in numeric_fields:
            grand_total[field] = sum(dept.get(field, 0) for dept in department_totals.values())
        
        return grand_total
    
    @staticmethod
    def _write_department_total_row(ws, row: int, label: str, totals: Dict[str, Any], border, headers: List[str]):
        from openpyxl.styles import Font, Alignment, PatternFill
        
        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        total_mapping = {
            'Sr.No': '',
            'Name': label,
            'Division': '',
            'Job Title': '',
            'Account No': '',
            'Working Days': '',
            'Basic Salary': totals.get('total_basic_salary', 0),
            'Transport Allowance': totals.get('total_transport', 0),
            'Meal Allowance': totals.get('total_meal', 0),
            'Telephone Allowance': totals.get('total_telephone', 0),
            'Fuel Allowance': totals.get('total_fuel', 0),
            'Attendance Bonus': totals.get('total_attendance_bonus', 0),
            'Performance Bonus': totals.get('total_performance_bonus', 0),
            'Regular Overtime': totals.get('total_regular_ot', 0),
            'Weekend Overtime': totals.get('total_weekend_ot', 0),
            'Total Overtime Hours': totals.get('total_ot_hours', 0),
            'Absent Days': totals.get('total_absent_days', 0),
            'Absent Deduction': totals.get('total_absent_deduction', 0),
            'Half Days': totals.get('total_half_days', 0),
            'Half Day Deduction': totals.get('total_half_day_deduction', 0),
            'Late Penalties': totals.get('total_late_penalty', 0),
            'Lunch Violations': totals.get('total_lunch_violations', 0),
            'Leave Deduction': totals.get('total_leave_deduction', 0),
            'Advance Deduction': totals.get('total_advance_deduction', 0),
            'EPF Deduction': totals.get('total_epf_deduction', 0),
            'Income Tax': totals.get('total_income_tax', 0),
            'Gross Salary': totals.get('total_gross', 0),
            'Total Deductions': totals.get('total_deductions', 0),
            'Net Salary': totals.get('total_net_salary', 0)
        }
        
        for col, header in enumerate(headers, 1):
            value = total_mapping.get(header, '')
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if isinstance(value, (int, float)) and col > 5:
                cell.number_format = '#,##0.00'
    
    @staticmethod
    def _format_excel_columns(ws, headers: List[str]):
        from openpyxl.utils import get_column_letter
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        ws.row_dimensions[1].height = 30

class PayrollDataFormatter:
    @staticmethod
    def format_currency(amount: Decimal) -> str:
        return f"LKR {amount:,.2f}"

    @staticmethod
    def format_percentage(value: Decimal) -> str:
        return f"{value:.2f}%"

    @staticmethod
    def format_hours(hours: Decimal) -> str:
        return f"{hours:.2f} hrs"

    @staticmethod
    def format_employee_summary(employee_data: Dict[str, Any]) -> Dict[str, str]:
        return {
            'employee_name': employee_data.get('employee_name', ''),
            'employee_code': employee_data.get('employee_code', ''),
            'department': employee_data.get('division', ''),
            'basic_salary': PayrollDataFormatter.format_currency(employee_data.get('basic_salary', Decimal('0.00'))),
            'gross_salary': PayrollDataFormatter.format_currency(employee_data.get('gross_salary', Decimal('0.00'))),
            'net_salary': PayrollDataFormatter.format_currency(employee_data.get('net_salary', Decimal('0.00'))),
            'working_days': str(employee_data.get('working_days', 0)),
            'overtime_hours': PayrollDataFormatter.format_hours(employee_data.get('overtime_hours', Decimal('0.00')))
        }


class PayrollPDFProcessor:
    @staticmethod
    def create_individual_payslip_pdf(
        employee_data: Dict[str, Any], year: int, month: int
    ) -> bytes:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        from reportlab.lib.units import inch
        import io

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        story = []
        story.extend(PayrollPDFProcessor._create_pdf_header(year, month))
        story.extend(PayrollPDFProcessor._create_employee_info_section(employee_data))
        story.extend(PayrollPDFProcessor._create_earnings_section(employee_data))
        story.extend(PayrollPDFProcessor._create_deductions_section(employee_data))
        story.extend(PayrollPDFProcessor._create_summary_section(employee_data))
        story.extend(PayrollPDFProcessor._create_pdf_footer())

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _create_pdf_header(year: int, month: int) -> List:
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import Paragraph, Spacer

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
        )

        title = Paragraph(f"PAYSLIP - {calendar.month_name[month]} {year}", title_style)
        return [title, Spacer(1, 12)]

    @staticmethod
    def _create_employee_info_section(employee_data: Dict[str, Any]) -> List:
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib.units import inch

        employee_info = [
            ["Employee Name:", employee_data.get("employee_name", "")],
            ["Employee Code:", employee_data.get("employee_code", "")],
            ["Department:", employee_data.get("division", "")],
            ["Job Title:", employee_data.get("job_title", "")],
            ["Account No:", employee_data.get("account_no", "")],
        ]

        info_table = Table(employee_info, colWidths=[2 * inch, 3 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("BACKGROUND", (1, 0), (1, -1), colors.white),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return [info_table, Spacer(1, 20)]

    @staticmethod
    def _create_earnings_section(employee_data: Dict[str, Any]) -> List:
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib.units import inch

        earnings_data = [
            ["EARNINGS", "AMOUNT (LKR)"],
            ["Basic Salary", f"{employee_data.get('basic_salary', 0):,.2f}"],
            ["Transport Allowance", f"{employee_data.get('transport_allowance', 0):,.2f}"],
            ["Meal Allowance", f"{employee_data.get('meal_allowance', 0):,.2f}"],
            ["Telephone Allowance", f"{employee_data.get('telephone_allowance', 0):,.2f}"],
            ["Fuel Allowance", f"{employee_data.get('fuel_allowance', 0):,.2f}"],
            ["Attendance Bonus", f"{employee_data.get('attendance_bonus', 0):,.2f}"],
            ["Performance Bonus", f"{employee_data.get('performance_bonus', 0):,.2f}"],
            ["Regular Overtime", f"{employee_data.get('regular_overtime_pay', 0):,.2f}"],
            ["Weekend Overtime", f"{employee_data.get('weekend_overtime_pay', 0):,.2f}"],
            ["GROSS SALARY", f"{employee_data.get('gross_salary', 0):,.2f}"],
        ]

        earnings_table = Table(earnings_data, colWidths=[3 * inch, 2 * inch])
        earnings_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.lightblue),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return [earnings_table, Spacer(1, 20)]

    @staticmethod
    def _create_deductions_section(employee_data: Dict[str, Any]) -> List:
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib.units import inch

        deductions_data = [
            ["DEDUCTIONS", "AMOUNT (LKR)"],
            ["Absent Deduction", f"{employee_data.get('absent_deduction', 0):,.2f}"],
            ["Half Day Deduction", f"{employee_data.get('half_day_deduction', 0):,.2f}"],
            ["Late Penalties", f"{employee_data.get('late_penalty', 0):,.2f}"],
            ["Lunch Violations", f"{employee_data.get('lunch_violation_penalty', 0):,.2f}"],
            ["Leave Deduction", f"{employee_data.get('leave_deduction', 0):,.2f}"],
            ["Advance Deduction", f"{employee_data.get('advance_deduction', 0):,.2f}"],
            ["EPF (8%)", f"{employee_data.get('epf_deduction', 0):,.2f}"],
            ["Income Tax", f"{employee_data.get('income_tax', 0):,.2f}"],
            ["TOTAL DEDUCTIONS", f"{employee_data.get('total_deductions', 0):,.2f}"],
        ]

        deductions_table = Table(deductions_data, colWidths=[3 * inch, 2 * inch])
        deductions_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.lightcoral),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return [deductions_table, Spacer(1, 20)]

    @staticmethod
    def _create_summary_section(employee_data: Dict[str, Any]) -> List:
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle, Spacer
        from reportlab.lib.units import inch

        summary_data = [
            ["SUMMARY", ""],
            ["Gross Salary", f"LKR {employee_data.get('gross_salary', 0):,.2f}"],
            ["Total Deductions", f"LKR {employee_data.get('total_deductions', 0):,.2f}"],
            ["NET SALARY", f"LKR {employee_data.get('net_salary', 0):,.2f}"],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.lightgreen),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        return [summary_table, Spacer(1, 20)]

    @staticmethod
    def _create_pdf_footer() -> List:
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, Spacer

        styles = getSampleStyleSheet()

        footer_text = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, styles["Normal"])

        signature_section = [
            Spacer(1, 30),
            Paragraph("Employee Signature: _____________________", styles["Normal"]),
            Spacer(1, 10),
            Paragraph("HR Signature: _____________________", styles["Normal"]),
        ]

        return [footer, Spacer(1, 20)] + signature_section


class PayrollUtilityHelper:
    @staticmethod
    def generate_payroll_reference_number(
        year: int, month: int, employee_code: str
    ) -> str:
        timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
        return f"PAY{year}{month:02d}{employee_code}{timestamp[-4:]}"

    @staticmethod
    def get_payroll_processing_date() -> date:
        processing_day = SystemConfiguration.get_int_setting(
            "PAYROLL_PROCESSING_DAY", 25
        )
        today = timezone.now().date()

        if today.day >= processing_day:
            return date(today.year, today.month, processing_day)
        else:
            if today.month == 1:
                return date(today.year - 1, 12, processing_day)
            else:
                return date(today.year, today.month - 1, processing_day)

    @staticmethod
    def calculate_working_days_between_dates(start_date: date, end_date: date) -> int:
        working_days = 0
        current_date = start_date

        while current_date <= end_date:
            if current_date.weekday() < 5 and not Holiday.is_holiday_date(current_date):
                working_days += 1
            current_date += timedelta(days=1)

        return working_days

    @staticmethod
    def get_next_payroll_period() -> Tuple[int, int]:
        today = timezone.now().date()
        processing_day = SystemConfiguration.get_int_setting(
            "PAYROLL_PROCESSING_DAY", 25
        )

        if today.day < processing_day:
            return today.year, today.month
        else:
            if today.month == 12:
                return today.year + 1, 1
            else:
                return today.year, today.month + 1

    @staticmethod
    def validate_payroll_amount(amount: Decimal, field_name: str) -> Tuple[bool, str]:
        if amount < 0:
            return False, f"{field_name} cannot be negative"

        max_amount = Decimal(
            SystemConfiguration.get_setting("MAX_PAYROLL_AMOUNT", "1000000.00")
        )
        if amount > max_amount:
            return False, f"{field_name} exceeds maximum allowed amount"

        return True, "Valid amount"

    @staticmethod
    def round_payroll_amount(amount: Decimal) -> Decimal:
        rounding_precision = SystemConfiguration.get_int_setting(
            "PAYROLL_ROUNDING_PRECISION", 2
        )
        return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    @staticmethod
    def get_payroll_backup_data(payroll_data: Dict[str, Any]) -> Dict[str, Any]:
        backup_data = {
            "timestamp": timezone.now().isoformat(),
            "payroll_data": payroll_data,
            "system_settings": PayrollUtilityHelper._get_relevant_system_settings(),
            "checksum": PayrollUtilityHelper._calculate_data_checksum(payroll_data),
        }
        return backup_data

    @staticmethod
    def _get_relevant_system_settings() -> Dict[str, str]:
        relevant_settings = [
            "NET_WORKING_HOURS",
            "OVERTIME_RATE_MULTIPLIER",
            "WEEKEND_OVERTIME_MULTIPLIER",
            "EPF_EMPLOYEE_RATE",
            "EPF_EMPLOYER_RATE",
            "ETF_RATE",
            "BASIC_TAX_RATE",
            "HALF_DAY_SALARY_PERCENTAGE",
            "ATTENDANCE_BONUS_THRESHOLD",
            "PUNCTUALITY_BONUS_THRESHOLD",
            "OTHER_STAFF_GRACE_PERIOD_MINUTES",
            "HALF_DAY_THRESHOLD_MINUTES",
            "LUNCH_VIOLATION_LIMIT_PER_MONTH",
            "MAX_LUNCH_DURATION_MINUTES",
            "SALARY_ADVANCE_MAX_PERCENTAGE",
        ]

        settings = {}
        for setting in relevant_settings:
            settings[setting] = SystemConfiguration.get_setting(setting, "")

        return settings

    @staticmethod
    def _calculate_data_checksum(data: Dict[str, Any]) -> str:
        import hashlib

        data_string = json.dumps(data, sort_keys=True, default=str)
        return hashlib.md5(data_string.encode()).hexdigest()


class PayrollCacheManager:
    @staticmethod
    def get_cache_key(prefix: str, *args) -> str:
        import hashlib

        key_parts = [prefix] + [str(arg) for arg in args]
        key_string = "_".join(key_parts)
        return hashlib.md5(key_string.encode()).hexdigest()

    @staticmethod
    def cache_payroll_calculation(
        employee_id: int,
        year: int,
        month: int,
        calculation_data: Dict[str, Any],
        timeout: int = 3600,
    ):
        from django.core.cache import cache

        cache_key = PayrollCacheManager.get_cache_key(
            "payroll_calc", employee_id, year, month
        )
        cache.set(cache_key, calculation_data, timeout)

    @staticmethod
    def get_cached_payroll_calculation(
        employee_id: int, year: int, month: int
    ) -> Optional[Dict[str, Any]]:
        from django.core.cache import cache

        cache_key = PayrollCacheManager.get_cache_key(
            "payroll_calc", employee_id, year, month
        )
        return cache.get(cache_key)

    @staticmethod
    def invalidate_payroll_cache(employee_id: int, year: int, month: int):
        from django.core.cache import cache

        cache_key = PayrollCacheManager.get_cache_key(
            "payroll_calc", employee_id, year, month
        )
        cache.delete(cache_key)

    @staticmethod
    def cache_department_summary(
        department_id: int, year: int, month: int, summary_data: Dict[str, Any]
    ):
        from django.core.cache import cache

        cache_key = PayrollCacheManager.get_cache_key(
            "dept_summary", department_id, year, month
        )
        cache.set(cache_key, summary_data, 7200)

    @staticmethod
    def get_cached_department_summary(
        department_id: int, year: int, month: int
    ) -> Optional[Dict[str, Any]]:
        from django.core.cache import cache

        cache_key = PayrollCacheManager.get_cache_key(
            "dept_summary", department_id, year, month
        )
        return cache.get(cache_key)


def get_current_payroll_period() -> Tuple[int, int]:
    today = timezone.now().date()
    processing_day = SystemConfiguration.get_int_setting("PAYROLL_PROCESSING_DAY", 25)

    if today.day >= processing_day:
        return today.year, today.month
    else:
        if today.month == 1:
            return today.year - 1, 12
        else:
            return today.year, today.month - 1


def safe_payroll_calculation(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            logger.error(f"Payroll calculation error in {func.__name__}: {str(e)}")
            return None

    return wrapper

def log_payroll_activity(user, action, details):
    from accounts.models import AuditLog
    try:
        AuditLog.objects.create(
            user=user,
            action=action,
            ip_address=getattr(user, 'last_login_ip', '') if user else '',
            user_agent=''
        )
    except Exception as e:
        logger.error(f"Failed to log payroll activity: {str(e)}")

