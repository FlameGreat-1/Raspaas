from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.db.models import Sum, Count, Avg
from django.contrib.admin import SimpleListFilter
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from django.core.exceptions import ValidationError
from accounts.models import CustomUser, Department, Role
from .models import (
    PayrollPeriod,
    Payslip,
    PayslipItem,
    SalaryAdvance,
    PayrollDepartmentSummary,
    PayrollBankTransfer,
)
from .utils import (
    PayrollExcelProcessor,
    PayrollPDFProcessor,
    PayrollReportDataProcessor,
    PayrollUtilityHelper,
    PayrollCacheManager,
    PayrollCalculator,
    PayrollDeductionCalculator,
    PayrollTaxCalculator,
    PayrollAdvanceCalculator,
    safe_payroll_calculation,
    log_payroll_activity,
)
from decimal import Decimal
import zipfile
import io


class PayrollPeriodStatusFilter(SimpleListFilter):
    title = "Status"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return PayrollPeriod.STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


class PayrollPeriodYearFilter(SimpleListFilter):
    title = "Year"
    parameter_name = "year"

    def lookups(self, request, model_admin):
        years = (
            PayrollPeriod.objects.values_list("year", flat=True)
            .distinct()
            .order_by("-year")
        )
        return [(year, str(year)) for year in years]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(year=self.value())
        return queryset


class PayslipStatusFilter(SimpleListFilter):
    title = "Status"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return Payslip.STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


class PayslipRoleFilter(SimpleListFilter):
    title = "Employee Role"
    parameter_name = "employee_role"

    def lookups(self, request, model_admin):
        roles = Role.objects.filter(is_active=True).values_list("id", "name")
        return [(role_id, role_name) for role_id, role_name in roles]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(employee__role_id=self.value())
        return queryset


class PayslipDepartmentFilter(SimpleListFilter):
    title = "Department"
    parameter_name = "department"

    def lookups(self, request, model_admin):
        departments = Department.objects.filter(is_active=True).values_list(
            "id", "name"
        )
        return [(dept_id, dept_name) for dept_id, dept_name in departments]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(employee__department_id=self.value())
        return queryset


class SalaryAdvanceStatusFilter(SimpleListFilter):
    title = "Status"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return SalaryAdvance.STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset


class SalaryAdvanceTypeFilter(SimpleListFilter):
    title = "Advance Type"
    parameter_name = "advance_type"

    def lookups(self, request, model_admin):
        return SalaryAdvance.ADVANCE_TYPES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(advance_type=self.value())
        return queryset


class DepartmentSummaryYearFilter(SimpleListFilter):
    title = "Year"
    parameter_name = "year"

    def lookups(self, request, model_admin):
        years = (
            PayrollDepartmentSummary.objects.values_list(
                "payroll_period__year", flat=True
            )
            .distinct()
            .order_by("-payroll_period__year")
        )
        return [(year, str(year)) for year in years]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(payroll_period__year=self.value())
        return queryset


class BankTransferStatusFilter(SimpleListFilter):
    title = "Status"
    parameter_name = "status"

    def lookups(self, request, model_admin):
        return PayrollBankTransfer.STATUS_CHOICES

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(status=self.value())
        return queryset
class PayrollPeriodAdmin(admin.ModelAdmin):
    list_display = [
        "period_display",
        "status_badge",
        "total_employees",
        "total_gross_display",
        "total_net_display",
        "processing_progress",
        "created_by",
        "created_at",
    ]
    list_filter = [PayrollPeriodStatusFilter, PayrollPeriodYearFilter, "created_at"]
    search_fields = ["year", "month"]

    fieldsets = (
        ("Period Information", {"fields": ("year", "month", "period_name", "status")}),
        (
            "Dates",
            {"fields": ("start_date", "end_date", "processing_date", "cutoff_date")},
        ),
        (
            "Financial Summary",
            {
                "fields": (
                    "total_employees",
                    "total_working_days",
                    "total_gross_salary",
                    "total_deductions",
                    "total_net_salary",
                )
            },
        ),
        (
            "Contributions",
            {
                "fields": (
                    "total_epf_employee",
                    "total_epf_employer",
                    "total_etf_contribution",
                )
            },
        ),
        (
            "Analytics",
            {
                "fields": ("role_based_summary", "department_summary"),
                "classes": ("collapse",),
            },
        ),
        (
            "Audit Information",
            {
                "fields": (
                    "created_by",
                    "approved_by",
                    "created_at",
                    "updated_at",
                    "approved_at",
                ),
                "classes": ("collapse",),
            },
        ),
    )
    actions = [
        "start_processing",
        "complete_processing",
        "approve_periods",
        "export_professional_excel",
        "create_payroll_backup",
    ]
    ordering = ["-year", "-month"]

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return [
                "id",
                "period_name",
                "total_employees",
                "total_working_days",
                "total_gross_salary",
                "total_deductions",
                "total_net_salary",
                "total_epf_employee",
                "total_epf_employer",
                "total_etf_contribution",
                "role_based_summary",
                "department_summary",
                "created_at",
                "updated_at",
                "approved_at",
            ]
        else:
            return [
                "id",
                "period_name",
                "total_employees",
                "total_working_days",
                "total_gross_salary",
                "total_deductions",
                "total_net_salary",
                "total_epf_employee",
                "total_epf_employer",
                "total_etf_contribution",
                "role_based_summary",
                "department_summary",
                "created_at",
                "updated_at",
                "approved_at",
            ]

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("created_by", "approved_by")

    def period_display(self, obj):
        return f"{obj.year}-{obj.month:02d} ({obj.period_name})"

    period_display.short_description = "Period"

    def status_badge(self, obj):
        colors = {
            "DRAFT": "#6c757d",
            "PROCESSING": "#007bff",
            "COMPLETED": "#28a745",
            "APPROVED": "#17a2b8",
            "PAID": "#28a745",
            "CANCELLED": "#dc3545",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 3px 8px; border-radius: 3px; font-size: 11px;">{}</span>',
            colors.get(obj.status, "#6c757d"),
            obj.get_status_display(),
        )

    status_badge.short_description = "Status"

    def total_gross_display(self, obj):
        return f"LKR {obj.total_gross_salary:,.2f}"

    total_gross_display.short_description = "Total Gross"

    def total_net_display(self, obj):
        return f"LKR {obj.total_net_salary:,.2f}"

    total_net_display.short_description = "Total Net"

    def processing_progress(self, obj):
        total = obj.payslips.count()
        calculated = obj.payslips.filter(status__in=["CALCULATED", "APPROVED"]).count()
        percentage = (calculated / total * 100) if total > 0 else 0
        return format_html(
            '<div style="width: 100px; background-color: #e9ecef; border-radius: 3px;"><div style="width: {}%; background-color: #007bff; height: 20px; border-radius: 3px; text-align: center; color: white; font-size: 11px; line-height: 20px;">{}%</div></div>',
            percentage,
            int(percentage),
        )

    processing_progress.short_description = "Progress"

    def start_processing(self, request, queryset):
        for period in queryset:
            try:
                if period.status == "DRAFT":
                    period.status = "PROCESSING"
                    period.save()
                    messages.success(request, f"Started processing {period.period_name}")
                else:
                    messages.warning(request, f"Cannot start processing {period.period_name} - not in draft status")
            except Exception as e:
                messages.error(request, f"Error processing {period.period_name}: {str(e)}")
    start_processing.short_description = "Start processing selected periods"

    def complete_processing(self, request, queryset):
        for period in queryset:
            try:
                if period.status == "PROCESSING":
                    period.status = "COMPLETED"
                    period.calculate_period_totals()
                    period.save()
                    messages.success(request, f"Completed processing {period.period_name}")
                else:
                    messages.warning(request, f"Cannot complete {period.period_name} - not in processing status")
            except Exception as e:
                messages.error(request, f"Error completing {period.period_name}: {str(e)}")
    complete_processing.short_description = "Complete processing selected periods"

    def approve_periods(self, request, queryset):
        for period in queryset:
            try:
                if period.status == "COMPLETED":
                    period.status = "APPROVED"
                    period.approved_by = request.user
                    period.approved_at = timezone.now()
                    period.save()
                    messages.success(request, f"Approved {period.period_name}")
                else:
                    messages.warning(request, f"Cannot approve {period.period_name} - not completed")
            except Exception as e:
                messages.error(request, f"Error approving {period.period_name}: {str(e)}")
    approve_periods.short_description = "Approve selected periods"

    def export_professional_excel(self, request, queryset):
        for period in queryset:
            payslips = period.payslips.filter(status__in=["CALCULATED", "APPROVED", "PAID"])

            payroll_data = []
            for payslip in payslips:
                payslip_dict = self._convert_payslip_to_dict(payslip)
                formatted_data = PayrollReportDataProcessor.prepare_individual_payslip_data(
                    payslip.employee, payslip_dict
                )
                payroll_data.append(formatted_data)

            excel_bytes = PayrollExcelProcessor.create_payroll_excel(
                payroll_data, period.year, period.month
            )

            response = HttpResponse(
                excel_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="payroll_{period.year}_{period.month:02d}_complete.xlsx"'
            return response
    export_professional_excel.short_description = "Export Professional Excel"

    def create_payroll_backup(self, request, queryset):
        import json

        for period in queryset:
            payslips_data = []
            for payslip in period.payslips.all():
                payslip_dict = self._convert_payslip_to_dict(payslip)
                payslips_data.append(payslip_dict)

            backup_data = PayrollUtilityHelper.get_payroll_backup_data({
                "period_id": str(period.id),
                "year": period.year,
                "month": period.month,
                "payslips": payslips_data,
            })

            response = HttpResponse(
                json.dumps(backup_data, indent=2, default=str),
                content_type="application/json",
            )
            response["Content-Disposition"] = f'attachment; filename="payroll_backup_{period.year}_{period.month:02d}.json"'
            return response
    create_payroll_backup.short_description = "Create Payroll Backup"

    def _convert_payslip_to_dict(self, payslip):
        from employees.models import EmployeeProfile

        try:
            profile = EmployeeProfile.objects.get(user=payslip.employee)
        except EmployeeProfile.DoesNotExist:
            profile = None

        return {
            "sr_no": payslip.sr_no,
            "employee_name": payslip.employee.get_full_name(),
            "employee_code": payslip.employee.employee_code,
            "division": payslip.employee.department.name if payslip.employee.department else "",
            "job_title": payslip.employee.job_title or "",
            "account_no": profile.bank_account_number if profile else "",
            "working_days": payslip.working_days,
            "basic_salary": payslip.basic_salary,
            "transport_allowance": payslip.transport_allowance,
            "telephone_allowance": payslip.telephone_allowance,
            "fuel_allowance": payslip.fuel_allowance,
            "meal_allowance": payslip.meal_allowance,
            "attendance_bonus": payslip.attendance_bonus,
            "performance_bonus": payslip.performance_bonus,
            "interim_allowance": payslip.interim_allowance,
            "education_allowance": payslip.education_allowance,
            "religious_pay": payslip.religious_pay,
            "friday_salary": payslip.friday_salary,
            "friday_overtime": payslip.friday_overtime,
            "regular_overtime_pay": payslip.regular_overtime,
            "weekend_overtime_pay": payslip.friday_overtime,
            "gross_salary": payslip.gross_salary,
            "working_day_meals": payslip.working_day_meals,
            "total_overtime_hours": payslip.overtime_hours,
            "leave_days": payslip.leave_days,
            "absent_days": payslip.leave_days,
            "half_days": 0,
            "absent_deduction": payslip.leave_deduction,
            "half_day_deduction": Decimal("0.00"),
            "leave_deduction": payslip.leave_deduction,
            "late_penalty": payslip.late_penalty,
            "lunch_violation_penalty": payslip.lunch_violation_penalty,
            "advance_deduction": payslip.advance_deduction,
            "epf_deduction": payslip.employee_epf_contribution,
            "income_tax": payslip.income_tax,
            "total_deductions": payslip.total_deductions,
            "net_salary": payslip.net_salary,
            "epf_salary_base": payslip.epf_salary_base,
        }
class PayslipAdmin(admin.ModelAdmin):
    list_display = [
        "sr_no",
        "employee_name_display",
        "employee_code_display",
        "division_display",
        "job_title_display",
        "account_no_display",
        "working_days",
        "basic_salary_display",
        "transport_allowance_display",
        "telephone_allowance_display",
        "fuel_allowance_display",
        "meal_allowance_display",
        "attendance_bonus_display",
        "performance_bonus_display",
        "interim_allowance_display",
        "education_allowance_display",
        "religious_pay_display",
        "friday_salary_display",
        "friday_overtime_display",
        "regular_overtime_display",
        "weekend_overtime_display",
        "gross_salary_display",
        "working_day_meals",
        "total_overtime_hours_display",
        "leave_days",
        "absent_days_display",
        "half_days_display",
        "absent_deduction_display",
        "half_day_deduction_display",
        "leave_deduction_display",
        "late_penalty_display",
        "lunch_violation_penalty_display",
        "advance_deduction_display",
        "epf_deduction_display",
        "income_tax_display",
        "total_deductions_display",
        "net_salary_display",
        "epf_salary_base_display",
        "status_badge",
    ]

    list_filter = [
        PayslipStatusFilter,
        PayslipRoleFilter,
        PayslipDepartmentFilter,
        "payroll_period__year",
        "payroll_period__month",
        "created_at",
    ]

    search_fields = [
        "sr_no",
        "employee__employee_code",
        "employee__first_name",
        "employee__last_name",
        "employee__email",
    ]

    readonly_fields = [
        "id",
        "reference_number",
        "created_at",
        "updated_at",
        "approved_at",
    ]

    actions = [
        "calculate_payslips",
        "approve_payslips",
        "export_professional_excel",
        "generate_individual_payslip_pdf",
        "validate_payroll_amounts",
        "generate_reference_numbers",
    ]

    ordering = [
        "-payroll_period__year",
        "-payroll_period__month",
        "sr_no",
    ]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related(
                "employee",
                "employee__role",
                "employee__department",
                "payroll_period",
                "calculated_by",
                "approved_by",
                "monthly_summary",
            )
            .prefetch_related("employee__employee_profile")  
        )

    def employee_name_display(self, obj):
        return obj.employee.get_full_name()

    employee_name_display.short_description = "Employee Name"

    def employee_code_display(self, obj):
        return obj.employee.employee_code

    employee_code_display.short_description = "Employee Code"

    def division_display(self, obj):
        return obj.employee.department.name if obj.employee.department else ""

    division_display.short_description = "Division"

    def job_title_display(self, obj):
        return obj.employee.job_title or ""

    job_title_display.short_description = "Job Title"

    def account_no_display(self, obj):
        try:
            from employees.models import EmployeeProfile

            profile = EmployeeProfile.objects.get(user=obj.employee)
            return profile.bank_account_number if profile else ""
        except:
            return ""

    account_no_display.short_description = "Account No"

    def basic_salary_display(self, obj):
        return f"LKR {obj.basic_salary:,.2f}"

    basic_salary_display.short_description = "Basic Salary"

    def transport_allowance_display(self, obj):
        return f"LKR {obj.transport_allowance:,.2f}"

    transport_allowance_display.short_description = "Transport Allowance"

    def telephone_allowance_display(self, obj):
        return f"LKR {obj.telephone_allowance:,.2f}"

    telephone_allowance_display.short_description = "Telephone Allowance"

    def fuel_allowance_display(self, obj):
        return f"LKR {obj.fuel_allowance:,.2f}"

    fuel_allowance_display.short_description = "Fuel Allowance"

    def meal_allowance_display(self, obj):
        return f"LKR {obj.meal_allowance:,.2f}"

    meal_allowance_display.short_description = "Meal Allowance"

    def attendance_bonus_display(self, obj):
        return f"LKR {obj.attendance_bonus:,.2f}"

    attendance_bonus_display.short_description = "Attendance Bonus"

    def performance_bonus_display(self, obj):
        return f"LKR {obj.performance_bonus:,.2f}"

    performance_bonus_display.short_description = "Performance Bonus"

    def interim_allowance_display(self, obj):
        return f"LKR {obj.interim_allowance:,.2f}"

    interim_allowance_display.short_description = "Interim Allowance"

    def education_allowance_display(self, obj):
        return f"LKR {obj.education_allowance:,.2f}"

    education_allowance_display.short_description = "Education Allowance"

    def religious_pay_display(self, obj):
        return f"LKR {obj.religious_pay:,.2f}"

    religious_pay_display.short_description = "Religious Pay"

    def friday_salary_display(self, obj):
        return f"LKR {obj.friday_salary:,.2f}"

    friday_salary_display.short_description = "Friday Salary"

    def friday_overtime_display(self, obj):
        return f"LKR {obj.friday_overtime:,.2f}"

    friday_overtime_display.short_description = "Friday Overtime"

    def regular_overtime_display(self, obj):
        return f"LKR {obj.regular_overtime:,.2f}"

    regular_overtime_display.short_description = "Regular Overtime"

    def weekend_overtime_display(self, obj):
        return f"LKR {obj.friday_overtime:,.2f}"

    weekend_overtime_display.short_description = "Weekend Overtime"

    def gross_salary_display(self, obj):
        return f"LKR {obj.gross_salary:,.2f}"

    gross_salary_display.short_description = "Gross Salary"

    def total_overtime_hours_display(self, obj):
        return f"{obj.overtime_hours:.2f} hrs"

    total_overtime_hours_display.short_description = "Total OT Hours"

    def absent_days_display(self, obj):
        return obj.leave_days

    absent_days_display.short_description = "Absent Days"

    def half_days_display(self, obj):
        return 0

    half_days_display.short_description = "Half Days"

    def absent_deduction_display(self, obj):
        return f"LKR {obj.leave_deduction:,.2f}"

    absent_deduction_display.short_description = "Absent Deduction"

    def half_day_deduction_display(self, obj):
        return "LKR 0.00"

    half_day_deduction_display.short_description = "Half Day Deduction"

    def leave_deduction_display(self, obj):
        return f"LKR {obj.leave_deduction:,.2f}"

    leave_deduction_display.short_description = "Leave Deduction"

    def late_penalty_display(self, obj):
        return f"LKR {obj.late_penalty:,.2f}"

    late_penalty_display.short_description = "Late Penalty"

    def lunch_violation_penalty_display(self, obj):
        return f"LKR {obj.lunch_violation_penalty:,.2f}"

    lunch_violation_penalty_display.short_description = "Lunch Violation"

    def advance_deduction_display(self, obj):
        return f"LKR {obj.advance_deduction:,.2f}"

    advance_deduction_display.short_description = "Advance Deduction"

    def epf_deduction_display(self, obj):
        return f"LKR {obj.employee_epf_contribution:,.2f}"

    epf_deduction_display.short_description = "EPF Deduction"

    def income_tax_display(self, obj):
        return f"LKR {obj.income_tax:,.2f}"

    income_tax_display.short_description = "Income Tax"

    def total_deductions_display(self, obj):
        return f"LKR {obj.total_deductions:,.2f}"

    total_deductions_display.short_description = "Total Deductions"

    def net_salary_display(self, obj):
        return f"LKR {obj.net_salary:,.2f}"

    net_salary_display.short_description = "Net Salary"

    def epf_salary_base_display(self, obj):
        return f"LKR {obj.epf_salary_base:,.2f}"

    epf_salary_base_display.short_description = "EPF Salary Base"

    def status_badge(self, obj):
        colors = {
            "DRAFT": "#6c757d",
            "CALCULATED": "#007bff",
            "APPROVED": "#28a745",
            "PAID": "#17a2b8",
            "CANCELLED": "#dc3545",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.status, "#6c757d"),
            obj.get_status_display(),
        )

    status_badge.short_description = "Status"

    def calculate_payslips(self, request, queryset):
        for payslip in queryset:
            try:
                if payslip.status == "DRAFT":
                    payslip.calculate_payroll()
                    payslip.calculated_by = request.user
                    payslip.save()
                    messages.success(
                        request,
                        f"Calculated payslip for {payslip.employee.get_full_name()}",
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot calculate {payslip.employee.get_full_name()} - not in draft status",
                    )
            except Exception as e:
                messages.error(
                    request,
                    f"Error calculating {payslip.employee.get_full_name()}: {str(e)}",
                )

    calculate_payslips.short_description = "Calculate selected payslips"

    def approve_payslips(self, request, queryset):
        for payslip in queryset:
            try:
                if payslip.status == "CALCULATED":
                    payslip.approve(request.user)
                    messages.success(
                        request,
                        f"Approved payslip for {payslip.employee.get_full_name()}",
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot approve {payslip.employee.get_full_name()} - not calculated",
                    )
            except Exception as e:
                messages.error(
                    request,
                    f"Error approving {payslip.employee.get_full_name()}: {str(e)}",
                )

    approve_payslips.short_description = "Approve selected payslips"

    def export_professional_excel(self, request, queryset):
        payroll_data = []
        for payslip in queryset:
            payslip_dict = self._convert_payslip_to_dict(payslip)
            formatted_data = PayrollReportDataProcessor.prepare_individual_payslip_data(
                payslip.employee, payslip_dict
            )
            payroll_data.append(formatted_data)

        year = queryset.first().payroll_period.year
        month = queryset.first().payroll_period.month
        excel_bytes = PayrollExcelProcessor.create_payroll_excel(
            payroll_data, year, month
        )

        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="payslips_{year}_{month:02d}_detailed.xlsx"'
        )
        return response

    export_professional_excel.short_description = "Export Professional Excel"

    def generate_individual_payslip_pdf(self, request, queryset):
        if queryset.count() == 1:
            payslip = queryset.first()
            payslip_dict = self._convert_payslip_to_dict(payslip)
            formatted_data = PayrollReportDataProcessor.prepare_individual_payslip_data(
                payslip.employee, payslip_dict
            )

            pdf_bytes = PayrollPDFProcessor.create_individual_payslip_pdf(
                formatted_data,
                payslip.payroll_period.year,
                payslip.payroll_period.month,
            )

            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="payslip_{payslip.employee.employee_code}.pdf"'
            )
            return response
        else:
            response = HttpResponse(content_type="application/zip")
            response["Content-Disposition"] = 'attachment; filename="payslips.zip"'

            with zipfile.ZipFile(response, "w") as zip_file:
                for payslip in queryset:
                    payslip_dict = self._convert_payslip_to_dict(payslip)
                    formatted_data = (
                        PayrollReportDataProcessor.prepare_individual_payslip_data(
                            payslip.employee, payslip_dict
                        )
                    )

                    pdf_bytes = PayrollPDFProcessor.create_individual_payslip_pdf(
                        formatted_data,
                        payslip.payroll_period.year,
                        payslip.payroll_period.month,
                    )

                    zip_file.writestr(
                        f"payslip_{payslip.employee.employee_code}.pdf", pdf_bytes
                    )

            return response

    generate_individual_payslip_pdf.short_description = "Generate PDF Payslips"

    def validate_payroll_amounts(self, request, queryset):
        validation_errors = []

        for payslip in queryset:
            fields_to_validate = [
                ("basic_salary", payslip.basic_salary),
                ("gross_salary", payslip.gross_salary),
                ("net_salary", payslip.net_salary),
                ("total_deductions", payslip.total_deductions),
            ]

            for field_name, amount in fields_to_validate:
                is_valid, message = PayrollUtilityHelper.validate_payroll_amount(
                    amount, field_name
                )
                if not is_valid:
                    validation_errors.append(
                        f"{payslip.employee.employee_code}: {message}"
                    )

        if validation_errors:
            for error in validation_errors:
                messages.error(request, error)
        else:
            messages.success(
                request, f"All {queryset.count()} payslips passed validation"
            )

    validate_payroll_amounts.short_description = "Validate Payroll Amounts"

    def generate_reference_numbers(self, request, queryset):
        for payslip in queryset:
            if not payslip.reference_number:
                payslip.reference_number = (
                    PayrollUtilityHelper.generate_payroll_reference_number(
                        payslip.payroll_period.year,
                        payslip.payroll_period.month,
                        payslip.employee.employee_code,
                    )
                )
                payslip.save()
                messages.success(
                    request, f"Generated reference: {payslip.reference_number}"
                )

    generate_reference_numbers.short_description = "Generate Reference Numbers"

    def _convert_payslip_to_dict(self, payslip):
        from employees.models import EmployeeProfile

        try:
            profile = EmployeeProfile.objects.get(user=payslip.employee)
        except EmployeeProfile.DoesNotExist:
            profile = None

        return {
            "sr_no": payslip.sr_no,
            "employee_name": payslip.employee.get_full_name(),
            "employee_code": payslip.employee.employee_code,
            "division": (
                payslip.employee.department.name if payslip.employee.department else ""
            ),
            "job_title": payslip.employee.job_title or "",
            "account_no": profile.bank_account_number if profile else "",
            "working_days": payslip.working_days,
            "basic_salary": payslip.basic_salary,
            "transport_allowance": payslip.transport_allowance,
            "telephone_allowance": payslip.telephone_allowance,
            "fuel_allowance": payslip.fuel_allowance,
            "meal_allowance": payslip.meal_allowance,
            "attendance_bonus": payslip.attendance_bonus,
            "performance_bonus": payslip.performance_bonus,
            "interim_allowance": payslip.interim_allowance,
            "education_allowance": payslip.education_allowance,
            "religious_pay": payslip.religious_pay,
            "friday_salary": payslip.friday_salary,
            "friday_overtime": payslip.friday_overtime,
            "regular_overtime_pay": payslip.regular_overtime,
            "weekend_overtime_pay": payslip.friday_overtime,
            "gross_salary": payslip.gross_salary,
            "working_day_meals": payslip.working_day_meals,
            "total_overtime_hours": payslip.overtime_hours,
            "leave_days": payslip.leave_days,
            "absent_days": payslip.leave_days,
            "half_days": 0,
            "absent_deduction": payslip.leave_deduction,
            "half_day_deduction": Decimal("0.00"),
            "leave_deduction": payslip.leave_deduction,
            "late_penalty": payslip.late_penalty,
            "lunch_violation_penalty": payslip.lunch_violation_penalty,
            "advance_deduction": payslip.advance_deduction,
            "epf_deduction": payslip.employee_epf_contribution,
            "income_tax": payslip.income_tax,
            "total_deductions": payslip.total_deductions,
            "net_salary": payslip.net_salary,
            "epf_salary_base": payslip.epf_salary_base,
        }
class PayslipItemAdmin(admin.ModelAdmin):
    list_display = [
        "payslip_info",
        "item_type_badge",
        "item_name",
        "amount_display",
        "rate",
        "quantity",
        "is_taxable",
        "is_epf_applicable",
    ]
    list_filter = [
        "item_type",
        "is_taxable",
        "is_epf_applicable",
        "is_mandatory",
        "payslip__payroll_period__year",
        "payslip__payroll_period__month",
    ]
    search_fields = [
        "item_code",
        "item_name",
        "payslip__employee__employee_code",
        "payslip__employee__first_name",
        "payslip__employee__last_name",
    ]
    readonly_fields = ["id", "created_at", "updated_at"]
    fieldsets = (
        (
            "Item Information",
            {
                "fields": (
                    "payslip",
                    "item_type",
                    "item_code",
                    "item_name",
                    "description",
                )
            },
        ),
        (
            "Amount & Calculation",
            {"fields": ("amount", "rate", "quantity", "calculation_basis")},
        ),
        (
            "Tax & Contribution Settings",
            {"fields": ("is_taxable", "is_epf_applicable", "is_mandatory")},
        ),
        (
            "Calculation Data",
            {"fields": ("calculation_data",), "classes": ("collapse",)},
        ),
        (
            "Audit Information",
            {"fields": ("created_at", "updated_at"), "classes": ("collapse",)},
        ),
    )
    ordering = ["payslip", "item_type", "item_code"]

    def get_queryset(self, request):
        return (
            super().get_queryset(request).select_related("payslip", "payslip__employee")
        )

    def payslip_info(self, obj):
        return format_html(
            "<strong>{}</strong><br><small>{}</small>",
            obj.payslip.employee.get_full_name(),
            obj.payslip.reference_number,
        )

    payslip_info.short_description = "Payslip"

    def item_type_badge(self, obj):
        colors = {
            "EARNING": "#28a745",
            "DEDUCTION": "#dc3545",
            "CONTRIBUTION": "#007bff",
            "TAX": "#ffc107",
            "BONUS": "#17a2b8",
            "ALLOWANCE": "#6f42c1",
            "PENALTY": "#fd7e14",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.item_type, "#6c757d"),
            obj.get_item_type_display(),
        )

    item_type_badge.short_description = "Type"

    def amount_display(self, obj):
        return f"LKR {obj.amount:,.2f}"

    amount_display.short_description = "Amount"


class PayslipItemInline(admin.TabularInline):
    model = PayslipItem
    extra = 0
    fields = [
        "item_type",
        "item_name",
        "amount",
        "rate",
        "quantity",
        "is_taxable",
        "is_epf_applicable",
    ]
    readonly_fields = ["amount"]
class SalaryAdvanceAdmin(admin.ModelAdmin):
    list_display = [
        "reference_number",
        "employee_info",
        "advance_type_badge",
        "status_badge",
        "amount_display",
        "outstanding_display",
        "installments",
        "requested_date",
        "overdue_status",
    ]
    list_filter = [
        SalaryAdvanceStatusFilter,
        SalaryAdvanceTypeFilter,
        "requested_date",
        "approved_date",
    ]
    search_fields = [
        "reference_number",
        "employee__employee_code",
        "employee__first_name",
        "employee__last_name",
        "reason",
    ]
    readonly_fields = [
        "id",
        "reference_number",
        "employee_basic_salary",
        "max_allowed_percentage",
        "advance_count_this_year",
        "is_overdue",
        "created_at",
        "updated_at",
        "approved_date",
        "disbursement_date",
        "completion_date",
    ]
    fieldsets = (
        (
            "Basic Information",
            {"fields": ("employee", "advance_type", "reference_number", "status")},
        ),
        (
            "Financial Details",
            {
                "fields": (
                    "amount",
                    "outstanding_amount",
                    "monthly_deduction",
                    "installments",
                )
            },
        ),
        ("Request Information", {"fields": ("reason", "purpose_details")}),
        (
            "Employee Context",
            {
                "fields": (
                    "employee_basic_salary",
                    "max_allowed_percentage",
                    "advance_count_this_year",
                )
            },
        ),
        (
            "Timeline",
            {
                "fields": (
                    "requested_date",
                    "approved_date",
                    "disbursement_date",
                    "completion_date",
                )
            },
        ),
        ("Status Information", {"fields": ("is_overdue", "is_active")}),
        (
            "Audit Information",
            {
                "fields": ("requested_by", "approved_by", "created_at", "updated_at"),
                "classes": ("collapse",),
            },
        ),
    )
    actions = [
        "approve_advances",
        "activate_advances",
        "cancel_advances",
        "validate_advance_eligibility",
        "export_advances",
    ]
    ordering = ["-created_at"]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related(
                "employee",
                "employee__role",
                "employee__department",
                "requested_by",
                "approved_by",
            )
        )

    def employee_info(self, obj):
        return format_html(
            "<strong>{}</strong><br><small>{} - {}</small>",
            obj.employee.get_full_name(),
            obj.employee.employee_code,
            obj.employee.role.name if obj.employee.role else "No Role",
        )

    employee_info.short_description = "Employee"

    def advance_type_badge(self, obj):
        colors = {
            "SALARY": "#007bff",
            "EMERGENCY": "#dc3545",
            "PURCHASE": "#28a745",
            "MEDICAL": "#ffc107",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.advance_type, "#6c757d"),
            obj.get_advance_type_display(),
        )

    advance_type_badge.short_description = "Type"

    def status_badge(self, obj):
        colors = {
            "PENDING": "#ffc107",
            "APPROVED": "#17a2b8",
            "ACTIVE": "#28a745",
            "COMPLETED": "#6c757d",
            "CANCELLED": "#dc3545",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.status, "#6c757d"),
            obj.get_status_display(),
        )

    status_badge.short_description = "Status"

    def amount_display(self, obj):
        return f"LKR {obj.amount:,.2f}"

    amount_display.short_description = "Amount"

    def outstanding_display(self, obj):
        if obj.outstanding_amount > 0:
            return format_html(
                '<span style="color: red;">LKR {}</span>',
                f"{obj.outstanding_amount:,.2f}",
            )
        return format_html('<span style="color: green;">LKR 0.00</span>')

    outstanding_display.short_description = "Outstanding"

    def overdue_status(self, obj):
        if obj.is_overdue:
            return format_html(
                '<span style="color: red; font-weight: bold;">⚠ OVERDUE</span>'
            )
        elif obj.status == "ACTIVE":
            return format_html('<span style="color: green;">✓ Active</span>')
        return format_html('<span style="color: gray;">-</span>')

    overdue_status.short_description = "Overdue"

    def approve_advances(self, request, queryset):
        for advance in queryset:
            try:
                if advance.status == "PENDING":
                    advance.approve(request.user)
                    messages.success(
                        request, f"Approved advance {advance.reference_number}"
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot approve {advance.reference_number} - not pending",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error approving {advance.reference_number}: {str(e)}"
                )

    approve_advances.short_description = "Approve selected advances"

    def activate_advances(self, request, queryset):
        for advance in queryset:
            try:
                if advance.status == "APPROVED":
                    advance.activate(request.user)
                    messages.success(
                        request, f"Activated advance {advance.reference_number}"
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot activate {advance.reference_number} - not approved",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error activating {advance.reference_number}: {str(e)}"
                )

    activate_advances.short_description = "Activate selected advances"

    def cancel_advances(self, request, queryset):
        for advance in queryset:
            try:
                if advance.status in ["PENDING", "APPROVED"]:
                    advance.status = "CANCELLED"
                    advance.save()
                    messages.success(
                        request, f"Cancelled advance {advance.reference_number}"
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot cancel {advance.reference_number} - already processed",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error cancelling {advance.reference_number}: {str(e)}"
                )

    cancel_advances.short_description = "Cancel selected advances"

    def validate_advance_eligibility(self, request, queryset):
        for advance in queryset:
            availability = PayrollAdvanceCalculator.calculate_available_advance_amount(
                advance.employee
            )

            if advance.amount > availability["available_amount"]:
                messages.error(
                    request,
                    f"{advance.employee.employee_code}: Advance amount {advance.amount} "
                    f"exceeds available limit {availability['available_amount']}",
                )
            else:
                messages.success(
                    request,
                    f"{advance.employee.employee_code}: Advance eligible. "
                    f"Available: {availability['available_amount']}",
                )

    validate_advance_eligibility.short_description = "Validate Advance Eligibility"

    def export_advances(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="salary_advances.csv"'

        import csv

        writer = csv.writer(response)
        writer.writerow(
            [
                "Reference",
                "Employee",
                "Type",
                "Amount",
                "Outstanding",
                "Status",
                "Requested Date",
            ]
        )

        for advance in queryset:
            writer.writerow(
                [
                    advance.reference_number,
                    advance.employee.get_full_name(),
                    advance.get_advance_type_display(),
                    advance.amount,
                    advance.outstanding_amount,
                    advance.get_status_display(),
                    advance.requested_date,
                ]
            )

        return response

    export_advances.short_description = "Export selected advances to CSV"
class PayrollDepartmentSummaryAdmin(admin.ModelAdmin):
    list_display = [
        "department_period_display",
        "employee_count",
        "total_basic_salary_display",
        "total_allowances_display", 
        "total_overtime_pay_display",
        "total_gross_salary_display",
        "total_deductions_display",
        "total_net_salary_display",
        "total_epf_employee_display",
        "total_epf_employer_display", 
        "total_etf_contribution_display",
        "average_salary_display",
        "budget_utilization_display",
        "efficiency_score_display",
        "created_at",
    ]
    list_filter = [
        DepartmentSummaryYearFilter,
        "payroll_period__month",
        "department",
        "created_at",
    ]
    search_fields = [
        "department__name",
        "payroll_period__year",
        "payroll_period__month",
    ]
    readonly_fields = [
        "id",
        "employee_count",
        "total_basic_salary",
        "total_allowances",
        "total_overtime_pay",
        "total_gross_salary",
        "total_deductions",
        "total_net_salary",
        "total_epf_employee",
        "total_epf_employer",
        "total_etf_contribution",
        "average_salary",
        "budget_utilization_percentage",
        "role_breakdown",
        "performance_metrics",
        "created_at",
        "updated_at",
    ]
    fieldsets = (
        ("Summary Information", {"fields": ("payroll_period", "department")}),
        ("Employee Statistics", {"fields": ("employee_count", "average_salary")}),
        (
            "Financial Summary",
            {
                "fields": (
                    "total_basic_salary",
                    "total_allowances",
                    "total_overtime_pay",
                    "total_gross_salary",
                    "total_deductions",
                    "total_net_salary",
                )
            },
        ),
        (
            "Contributions",
            {
                "fields": (
                    "total_epf_employee",
                    "total_epf_employer",
                    "total_etf_contribution",
                )
            },
        ),
        (
            "Budget Analysis",
            {"fields": ("department_budget", "budget_utilization_percentage")},
        ),
        (
            "Analytics",
            {
                "fields": ("role_breakdown", "performance_metrics"),
                "classes": ("collapse",),
            },
        ),
        (
            "Audit Information",
            {"fields": ("created_at", "updated_at"), "classes": ("collapse",)},
        ),
    )
    actions = ["recalculate_summaries", "export_summaries"]
    ordering = ["-payroll_period__year", "-payroll_period__month", "department__name"]

    def get_queryset(self, request):
        return (
            super().get_queryset(request).select_related("payroll_period", "department")
        )

    def department_period_display(self, obj):
        return format_html(
            "<strong>{}</strong><br><small>{}-{:02d}</small>",
            obj.department.name,
            obj.payroll_period.year,
            obj.payroll_period.month,
        )
    department_period_display.short_description = "Department & Period"

    def total_basic_salary_display(self, obj):
        return f"LKR {obj.total_basic_salary:,.2f}"
    total_basic_salary_display.short_description = "Total Basic Salary"

    def total_allowances_display(self, obj):
        return f"LKR {obj.total_allowances:,.2f}"
    total_allowances_display.short_description = "Total Allowances"

    def total_overtime_pay_display(self, obj):
        return f"LKR {obj.total_overtime_pay:,.2f}"
    total_overtime_pay_display.short_description = "Total Overtime Pay"

    def total_gross_salary_display(self, obj):
        return f"LKR {obj.total_gross_salary:,.2f}"
    total_gross_salary_display.short_description = "Total Gross Salary"

    def total_deductions_display(self, obj):
        return f"LKR {obj.total_deductions:,.2f}"
    total_deductions_display.short_description = "Total Deductions"

    def total_net_salary_display(self, obj):
        return f"LKR {obj.total_net_salary:,.2f}"
    total_net_salary_display.short_description = "Total Net Salary"

    def total_epf_employee_display(self, obj):
        return f"LKR {obj.total_epf_employee:,.2f}"
    total_epf_employee_display.short_description = "Total EPF Employee"

    def total_epf_employer_display(self, obj):
        return f"LKR {obj.total_epf_employer:,.2f}"
    total_epf_employer_display.short_description = "Total EPF Employer"

    def total_etf_contribution_display(self, obj):
        return f"LKR {obj.total_etf_contribution:,.2f}"
    total_etf_contribution_display.short_description = "Total ETF Contribution"

    def average_salary_display(self, obj):
        return f"LKR {obj.average_salary:,.2f}"
    average_salary_display.short_description = "Average Salary"

    def budget_utilization_display(self, obj):
        percentage = obj.budget_utilization_percentage
        color = (
            "#28a745"
            if percentage <= 100
            else "#dc3545" if percentage > 110 else "#ffc107"
        )
        return format_html(
            '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
            color,
            percentage,
        )
    budget_utilization_display.short_description = "Budget Utilization"

    def efficiency_score_display(self, obj):
        if (
            obj.performance_metrics
            and "department_efficiency_score" in obj.performance_metrics
        ):
            score = obj.performance_metrics["department_efficiency_score"]
            color = (
                "#28a745" if score >= 80 else "#ffc107" if score >= 60 else "#dc3545"
            )
            return format_html(
                '<span style="color: {}; font-weight: bold;">{:.1f}</span>',
                color,
                score,
            )
        return format_html('<span style="color: gray;">-</span>')
    efficiency_score_display.short_description = "Efficiency Score"

    def recalculate_summaries(self, request, queryset):
        for summary in queryset:
            try:
                summary.calculate_summary()
                messages.success(
                    request, f"Recalculated summary for {summary.department.name}"
                )
            except Exception as e:
                messages.error(
                    request, f"Error recalculating {summary.department.name}: {str(e)}"
                )
    recalculate_summaries.short_description = "Recalculate selected summaries"

    def export_summaries(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            'attachment; filename="department_summaries.csv"'
        )

        import csv
        writer = csv.writer(response)
        writer.writerow(
            [
                "Department",
                "Period",
                "Employees",
                "Total Basic Salary",
                "Total Allowances",
                "Total Overtime Pay",
                "Total Gross Salary",
                "Total Deductions",
                "Total Net Salary",
                "Total EPF Employee",
                "Total EPF Employer",
                "Total ETF Contribution",
                "Average Salary",
                "Budget Utilization",
                "Efficiency Score",
            ]
        )

        for summary in queryset:
            efficiency_score = ""
            if (
                summary.performance_metrics
                and "department_efficiency_score" in summary.performance_metrics
            ):
                efficiency_score = f"{summary.performance_metrics['department_efficiency_score']:.1f}"

            writer.writerow(
                [
                    summary.department.name,
                    f"{summary.payroll_period.year}-{summary.payroll_period.month:02d}",
                    summary.employee_count,
                    summary.total_basic_salary,
                    summary.total_allowances,
                    summary.total_overtime_pay,
                    summary.total_gross_salary,
                    summary.total_deductions,
                    summary.total_net_salary,
                    summary.total_epf_employee,
                    summary.total_epf_employer,
                    summary.total_etf_contribution,
                    summary.average_salary,
                    f"{summary.budget_utilization_percentage:.1f}%",
                    efficiency_score,
                ]
            )

        return response
    export_summaries.short_description = "Export selected summaries to CSV"

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
class PayrollBankTransferAdmin(admin.ModelAdmin):
    list_display = [
        "batch_reference",
        "period_info",
        "status_badge",
        "total_employees",
        "total_amount_display",
        "file_status",
        "processing_timeline",
        "created_by",
        "created_at",
    ]
    list_filter = [
        BankTransferStatusFilter,
        "payroll_period__year",
        "payroll_period__month",
        "bank_file_format",
        "created_at",
    ]
    search_fields = ["batch_reference", "payroll_period__year", "payroll_period__month"]
    readonly_fields = [
        "id",
        "batch_reference",
        "total_employees",
        "total_amount",
        "bank_file_path",
        "generated_at",
        "sent_at",
        "processed_at",
        "created_at",
        "updated_at",
    ]
    fieldsets = (
        (
            "Transfer Information",
            {"fields": ("payroll_period", "batch_reference", "status")},
        ),
        ("Financial Summary", {"fields": ("total_employees", "total_amount")}),
        ("File Information", {"fields": ("bank_file_path", "bank_file_format")}),
        (
            "Processing Timeline",
            {"fields": ("generated_at", "sent_at", "processed_at")},
        ),
        (
            "Bank Response",
            {"fields": ("bank_response", "error_details"), "classes": ("collapse",)},
        ),
        (
            "Audit Information",
            {
                "fields": ("created_by", "created_at", "updated_at"),
                "classes": ("collapse",),
            },
        ),
    )
    actions = [
        "generate_transfer_files",
        "mark_as_sent",
        "mark_as_processed",
        "export_transfers",
    ]
    ordering = ["-created_at"]

    def get_queryset(self, request):
        return (
            super().get_queryset(request).select_related("payroll_period", "created_by")
        )

    def period_info(self, obj):
        return format_html(
            "<strong>{}</strong><br><small>{}</small>",
            obj.payroll_period.period_name,
            f"{obj.payroll_period.year}-{obj.payroll_period.month:02d}",
        )

    period_info.short_description = "Period"

    def status_badge(self, obj):
        colors = {
            "PENDING": "#6c757d",
            "GENERATED": "#007bff",
            "SENT": "#ffc107",
            "PROCESSED": "#17a2b8",
            "COMPLETED": "#28a745",
            "FAILED": "#dc3545",
        }
        return format_html(
            '<span style="background-color: {}; color: white; padding: 2px 6px; border-radius: 3px; font-size: 10px;">{}</span>',
            colors.get(obj.status, "#6c757d"),
            obj.get_status_display(),
        )

    status_badge.short_description = "Status"

    def total_amount_display(self, obj):
        return f"LKR {obj.total_amount:,.2f}"

    total_amount_display.short_description = "Total Amount"

    def file_status(self, obj):
        if obj.bank_file_path:
            return format_html(
                '<span style="color: green;">✓ Generated</span><br><small>{}</small>',
                obj.bank_file_format,
            )
        return format_html('<span style="color: gray;">Not Generated</span>')

    file_status.short_description = "File Status"

    def processing_timeline(self, obj):
        timeline = []
        if obj.generated_at:
            timeline.append(f"Generated: {obj.generated_at.strftime('%m/%d %H:%M')}")
        if obj.sent_at:
            timeline.append(f"Sent: {obj.sent_at.strftime('%m/%d %H:%M')}")
        if obj.processed_at:
            timeline.append(f"Processed: {obj.processed_at.strftime('%m/%d %H:%M')}")

        if timeline:
            return format_html("<br>".join(timeline))
        return format_html('<span style="color: gray;">Not Started</span>')

    processing_timeline.short_description = "Timeline"

    def generate_transfer_files(self, request, queryset):
        for transfer in queryset:
            try:
                if transfer.status == "PENDING":
                    file_path = transfer.generate_bank_file()
                    if file_path:
                        messages.success(
                            request, f"Generated file for {transfer.batch_reference}"
                        )
                    else:
                        messages.error(
                            request,
                            f"Failed to generate file for {transfer.batch_reference}",
                        )
                else:
                    messages.warning(
                        request,
                        f"Cannot generate file for {transfer.batch_reference} - not pending",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error generating {transfer.batch_reference}: {str(e)}"
                )

    generate_transfer_files.short_description = "Generate transfer files"

    def mark_as_sent(self, request, queryset):
        for transfer in queryset:
            try:
                if transfer.status == "GENERATED":
                    transfer.status = "SENT"
                    transfer.sent_at = timezone.now()
                    transfer.save()
                    messages.success(
                        request, f"Marked {transfer.batch_reference} as sent"
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot mark {transfer.batch_reference} as sent - not generated",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error updating {transfer.batch_reference}: {str(e)}"
                )

    mark_as_sent.short_description = "Mark as sent to bank"

    def mark_as_processed(self, request, queryset):
        for transfer in queryset:
            try:
                if transfer.status == "SENT":
                    transfer.status = "PROCESSED"
                    transfer.processed_at = timezone.now()
                    transfer.save()
                    messages.success(
                        request, f"Marked {transfer.batch_reference} as processed"
                    )
                else:
                    messages.warning(
                        request,
                        f"Cannot mark {transfer.batch_reference} as processed - not sent",
                    )
            except Exception as e:
                messages.error(
                    request, f"Error updating {transfer.batch_reference}: {str(e)}"
                )

    mark_as_processed.short_description = "Mark as processed by bank"

    def export_transfers(self, request, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="bank_transfers.csv"'

        import csv

        writer = csv.writer(response)
        writer.writerow(
            [
                "Batch Reference",
                "Period",
                "Status",
                "Employees",
                "Total Amount",
                "Generated At",
                "Sent At",
            ]
        )

        for transfer in queryset:
            writer.writerow(
                [
                    transfer.batch_reference,
                    f"{transfer.payroll_period.year}-{transfer.payroll_period.month:02d}",
                    transfer.get_status_display(),
                    transfer.total_employees,
                    transfer.total_amount,
                    (
                        transfer.generated_at.strftime("%Y-%m-%d %H:%M")
                        if transfer.generated_at
                        else ""
                    ),
                    (
                        transfer.sent_at.strftime("%Y-%m-%d %H:%M")
                        if transfer.sent_at
                        else ""
                    ),
                ]
            )

        return response

    export_transfers.short_description = "Export selected transfers to CSV"

    def has_delete_permission(self, request, obj=None):
        if obj and obj.status in ["SENT", "PROCESSED", "COMPLETED"]:
            return False
        return request.user.is_superuser


class PayslipInline(admin.TabularInline):
    model = Payslip
    extra = 0
    fields = ["employee", "status", "basic_salary", "gross_salary", "net_salary"]
    readonly_fields = ["employee", "basic_salary", "gross_salary", "net_salary"]
    can_delete = False
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False


class PayrollDepartmentSummaryInline(admin.TabularInline):
    model = PayrollDepartmentSummary
    extra = 0
    fields = [
        "department",
        "employee_count",
        "total_gross_salary",
        "total_net_salary",
        "average_salary",
    ]
    readonly_fields = [
        "department",
        "employee_count",
        "total_gross_salary",
        "total_net_salary",
        "average_salary",
    ]
    can_delete = False
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False


class PayrollBankTransferInline(admin.TabularInline):
    model = PayrollBankTransfer
    extra = 0
    fields = [
        "batch_reference",
        "status",
        "total_employees",
        "total_amount",
        "generated_at",
    ]
    readonly_fields = [
        "batch_reference",
        "total_employees",
        "total_amount",
        "generated_at",
    ]
    can_delete = False
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False


class SalaryAdvanceInline(admin.TabularInline):
    model = SalaryAdvance
    extra = 0
    fields = [
        "advance_type",
        "amount",
        "outstanding_amount",
        "status",
        "requested_date",
    ]
    readonly_fields = ["amount", "outstanding_amount", "requested_date"]
    can_delete = False
    show_change_link = True

    def has_add_permission(self, request, obj=None):
        return False


def export_to_excel(modeladmin, request, queryset):
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{modeladmin.model._meta.verbose_name_plural}.xlsx"'
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = modeladmin.model._meta.verbose_name_plural

    fields = [
        field.name
        for field in modeladmin.model._meta.fields
        if not field.name.endswith("_ptr")
    ]
    headers = [modeladmin.model._meta.get_field(field).verbose_name for field in fields]

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = getattr(obj, field)
            if hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, Decimal):
                value = float(value)
            worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(response)
    return response


export_to_excel.short_description = "Export selected items to Excel"


def mark_as_processed(modeladmin, request, queryset):
    updated = 0
    for obj in queryset:
        if hasattr(obj, "status") and obj.status in ["PENDING", "GENERATED", "SENT"]:
            if hasattr(obj, "mark_as_processed"):
                obj.mark_as_processed(request.user)
                updated += 1
            elif hasattr(obj, "status"):
                obj.status = "PROCESSED"
                obj.save()
                updated += 1

    messages.success(request, f"Marked {updated} items as processed")


mark_as_processed.short_description = "Mark selected items as processed"


def recalculate_totals(modeladmin, request, queryset):
    updated = 0
    for obj in queryset:
        try:
            if hasattr(obj, "calculate_totals"):
                obj.calculate_totals()
                updated += 1
            elif hasattr(obj, "calculate_summary"):
                obj.calculate_summary()
                updated += 1
            elif hasattr(obj, "calculate_period_totals"):
                obj.calculate_period_totals()
                updated += 1
        except Exception as e:
            messages.error(request, f"Error recalculating {obj}: {str(e)}")

    if updated > 0:
        messages.success(request, f"Recalculated totals for {updated} items")


recalculate_totals.short_description = "Recalculate totals for selected items"


class PayrollAdminMixin:
    def get_readonly_fields(self, request, obj=None):
        readonly_fields = list(super().get_readonly_fields(request, obj))

        if obj and hasattr(obj, "status"):
            if obj.status in ["APPROVED", "PAID", "COMPLETED"]:
                readonly_fields.extend(
                    [
                        field.name
                        for field in obj._meta.fields
                        if field.name not in readonly_fields
                        and field.name not in ["status", "updated_at"]
                    ]
                )

        return readonly_fields

    def get_actions(self, request):
        actions = super().get_actions(request)

        actions["export_to_excel"] = (
            export_to_excel,
            "export_to_excel",
            export_to_excel.short_description,
        )

        actions["mark_as_processed"] = (
            mark_as_processed,
            "mark_as_processed",
            mark_as_processed.short_description,
        )

        actions["recalculate_totals"] = (
            recalculate_totals,
            "recalculate_totals",
            recalculate_totals.short_description,
        )

        return actions

    def save_model(self, request, obj, form, change):
        if not change:
            if hasattr(obj, "created_by"):
                obj.created_by = request.user

        if hasattr(obj, "updated_by"):
            obj.updated_by = request.user

        super().save_model(request, obj, form, change)


class PayrollModelAdmin(PayrollAdminMixin, admin.ModelAdmin):
    pass


class PayrollTabularInline(PayrollAdminMixin, admin.TabularInline):
    pass


class PayrollStackedInline(PayrollAdminMixin, admin.StackedInline):
    pass


PayrollPeriodAdmin.inlines = [
    PayslipInline,
    PayrollDepartmentSummaryInline,
    PayrollBankTransferInline,
]

PayslipAdmin.inlines = [PayslipItemInline]


class PayrollDashboardAdmin(admin.ModelAdmin):
    change_list_template = "admin/payroll/dashboard.html"

    def changelist_view(self, request, extra_context=None):
        try:
            current_year = timezone.now().year
            current_month = timezone.now().month

            dashboard_data = {
                "active_employees": CustomUser.objects.filter(
                    is_active=True, status="ACTIVE"
                ).count(),
                "current_period_exists": PayrollPeriod.objects.filter(
                    year=current_year, month=current_month
                ).exists(),
                "pending_advances": SalaryAdvance.objects.filter(
                    status="PENDING"
                ).count(),
                "draft_payslips": Payslip.objects.filter(status="DRAFT").count(),
                "processing_periods": PayrollPeriod.objects.filter(
                    status="PROCESSING"
                ).count(),
                "recent_periods": PayrollPeriod.objects.order_by("-year", "-month")[:5],
                "failed_transfers": PayrollBankTransfer.objects.filter(
                    status="FAILED"
                ).count(),
            }

            extra_context = extra_context or {}
            extra_context.update(
                {
                    "dashboard_data": dashboard_data,
                    "title": "Payroll Dashboard",
                    "has_add_permission": False,
                    "has_change_permission": False,
                    "has_delete_permission": False,
                }
            )
        except Exception as e:
            messages.error(request, f"Error loading dashboard: {str(e)}")
            extra_context = {"error": str(e)}

        return super().changelist_view(request, extra_context=extra_context)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_authenticated

    def has_delete_permission(self, request, obj=None):
        return False


class PayrollAnalyticsAdmin(admin.ModelAdmin):
    change_list_template = "admin/payroll/analytics.html"

    def changelist_view(self, request, extra_context=None):
        try:
            months = int(request.GET.get("months", 12))
            current_date = timezone.now().date()

            analytics_data = {
                "selected_months": months,
                "payroll_trends": self._get_payroll_trends(months),
                "department_analysis": self._get_department_analysis(),
                "role_analysis": self._get_role_analysis(),
                "advance_analysis": self._get_advance_analysis(),
            }

            extra_context = extra_context or {}
            extra_context.update(
                {
                    "analytics_data": analytics_data,
                    "title": "Payroll Analytics",
                    "has_add_permission": False,
                    "has_change_permission": False,
                    "has_delete_permission": False,
                }
            )
        except Exception as e:
            messages.error(request, f"Error loading analytics: {str(e)}")
            extra_context = {"error": str(e)}

        return super().changelist_view(request, extra_context=extra_context)

    def _get_payroll_trends(self, months):
        from datetime import datetime, timedelta

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=months * 30)

        periods = PayrollPeriod.objects.filter(
            start_date__gte=start_date, status__in=["COMPLETED", "APPROVED", "PAID"]
        ).order_by("year", "month")

        return [
            {
                "period": f"{p.year}-{p.month:02d}",
                "total_gross": float(p.total_gross_salary),
                "total_net": float(p.total_net_salary),
                "employee_count": p.total_employees,
            }
            for p in periods
        ]

    def _get_department_analysis(self):
        departments = PayrollDepartmentSummary.objects.select_related(
            "department"
        ).order_by("-total_gross_salary")[:10]
        return [
            {
                "department": dept.department.name,
                "total_gross": float(dept.total_gross_salary),
                "employee_count": dept.employee_count,
                "average_salary": float(dept.average_salary),
            }
            for dept in departments
        ]

    def _get_role_analysis(self):
        from django.db.models import Avg, Count

        role_data = (
            Payslip.objects.filter(status__in=["CALCULATED", "APPROVED", "PAID"])
            .values("employee__role__name")
            .annotate(avg_salary=Avg("gross_salary"), count=Count("id"))
            .order_by("-avg_salary")
        )

        return [
            {
                "role": data["employee__role__name"] or "No Role",
                "average_salary": float(data["avg_salary"] or 0),
                "employee_count": data["count"],
            }
            for data in role_data
        ]

    def _get_advance_analysis(self):
        return {
            "total_active": SalaryAdvance.objects.filter(status="ACTIVE").count(),
            "total_pending": SalaryAdvance.objects.filter(status="PENDING").count(),
            "total_overdue": SalaryAdvance.objects.filter(
                status="ACTIVE", is_overdue=True
            ).count(),
            "total_outstanding": float(
                SalaryAdvance.objects.filter(status="ACTIVE").aggregate(
                    total=Sum("outstanding_amount")
                )["total"]
                or 0
            ),
        }

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_authenticated

    def has_delete_permission(self, request, obj=None):
        return False


class PayrollMaintenanceAdmin(admin.ModelAdmin):
    change_list_template = "admin/payroll/maintenance.html"

    def changelist_view(self, request, extra_context=None):
        if request.method == "POST":
            action = request.POST.get("action")

            try:
                if action == "cleanup_data":
                    days = int(request.POST.get("days", 365))
                    result = self._cleanup_expired_data(days)
                    messages.success(request, f"Cleanup completed: {result}")

                elif action == "validate_integrity":
                    result = self._validate_data_integrity()
                    if result["is_valid"]:
                        messages.success(request, "Data integrity validation passed")
                    else:
                        messages.warning(
                            request, f"Found {result['issues_found']} integrity issues"
                        )

                elif action == "recalculate_totals":
                    period_id = request.POST.get("period_id")
                    if period_id:
                        period = PayrollPeriod.objects.get(id=period_id)
                        period.calculate_period_totals()
                        messages.success(request, "Period totals recalculated")

            except Exception as e:
                messages.error(request, f"Maintenance action failed: {str(e)}")

        try:
            integrity_status = self._validate_data_integrity()
            recent_periods = PayrollPeriod.objects.order_by("-year", "-month")[:10]

            extra_context = extra_context or {}
            extra_context.update(
                {
                    "integrity_status": integrity_status,
                    "recent_periods": recent_periods,
                    "title": "Payroll Maintenance",
                    "has_add_permission": False,
                    "has_change_permission": False,
                    "has_delete_permission": False,
                }
            )
        except Exception as e:
            messages.error(request, f"Error loading maintenance data: {str(e)}")
            extra_context = {"error": str(e)}

        return super().changelist_view(request, extra_context=extra_context)

    def _cleanup_expired_data(self, days):
        from datetime import timedelta

        cutoff_date = timezone.now() - timedelta(days=days)

        deleted_count = 0
        deleted_count += Payslip.objects.filter(
            created_at__lt=cutoff_date, status="CANCELLED"
        ).delete()[0]

        deleted_count += SalaryAdvance.objects.filter(
            created_at__lt=cutoff_date, status="CANCELLED"
        ).delete()[0]

        return f"Deleted {deleted_count} expired records"

    def _validate_data_integrity(self):
        issues = []

        payslips_without_period = Payslip.objects.filter(
            payroll_period__isnull=True
        ).count()
        if payslips_without_period > 0:
            issues.append(f"{payslips_without_period} payslips without payroll period")

        negative_salaries = Payslip.objects.filter(net_salary__lt=0).count()
        if negative_salaries > 0:
            issues.append(f"{negative_salaries} payslips with negative net salary")

        advances_without_employee = SalaryAdvance.objects.filter(
            employee__isnull=True
        ).count()
        if advances_without_employee > 0:
            issues.append(f"{advances_without_employee} advances without employee")

        return {
            "is_valid": len(issues) == 0,
            "issues_found": len(issues),
            "issues": issues,
        }

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return False


class PayrollBulkOperationsAdmin(admin.ModelAdmin):
    change_list_template = "admin/payroll/bulk_operations.html"

    def changelist_view(self, request, extra_context=None):
        if request.method == "POST":
            operation = request.POST.get("operation")

            try:
                if operation == "bulk_calculate":
                    period_id = request.POST.get("period_id")
                    employee_ids = request.POST.getlist("employee_ids")

                    if period_id:
                        result = self._bulk_calculate_payslips(
                            period_id, employee_ids, request.user
                        )
                        messages.success(
                            request,
                            f"Calculated {result['successful']} payslips, {result['failed']} failed",
                        )

                elif operation == "bulk_approve":
                    payslip_ids = request.POST.getlist("payslip_ids")
                    result = self._bulk_approve_payslips(payslip_ids, request.user)
                    messages.success(
                        request,
                        f"Processed {result['successful']} payslips for approval",
                    )

                elif operation == "period_processing":
                    period_id = request.POST.get("period_id")
                    action = request.POST.get("period_action")

                    if period_id and action:
                        period = PayrollPeriod.objects.get(id=period_id)
                        if action == "start":
                            period.status = "PROCESSING"
                            period.save()
                            messages.success(request, "Started period processing")
                        elif action == "complete":
                            period.status = "COMPLETED"
                            period.calculate_period_totals()
                            period.save()
                            messages.success(request, "Completed period processing")
                        elif action == "approve":
                            period.status = "APPROVED"
                            period.approved_by = request.user
                            period.approved_at = timezone.now()
                            period.save()
                            messages.success(request, "Approved payroll period")

            except Exception as e:
                messages.error(request, f"Bulk operation failed: {str(e)}")

        try:
            draft_periods = PayrollPeriod.objects.filter(
                status__in=["DRAFT", "PROCESSING"]
            )
            draft_payslips = Payslip.objects.filter(
                status="DRAFT", payroll_period__status__in=["DRAFT", "PROCESSING"]
            )[:100]
            calculated_payslips = Payslip.objects.filter(
                status="CALCULATED",
                payroll_period__status__in=["PROCESSING", "COMPLETED"],
            )[:100]

            extra_context = extra_context or {}
            extra_context.update(
                {
                    "draft_periods": draft_periods,
                    "draft_payslips": draft_payslips,
                    "calculated_payslips": calculated_payslips,
                    "title": "Bulk Operations",
                    "has_add_permission": False,
                    "has_change_permission": False,
                    "has_delete_permission": False,
                }
            )
        except Exception as e:
            messages.error(request, f"Error loading bulk operations data: {str(e)}")
            extra_context = {"error": str(e)}

        return super().changelist_view(request, extra_context=extra_context)

    def _bulk_calculate_payslips(self, period_id, employee_ids, user):
        period = PayrollPeriod.objects.get(id=period_id)
        payslips = Payslip.objects.filter(payroll_period=period, status="DRAFT")

        if employee_ids:
            payslips = payslips.filter(employee_id__in=employee_ids)

        successful = 0
        failed = 0

        for payslip in payslips:
            try:
                payslip.calculate_payroll()
                payslip.calculated_by = user
                payslip.save()
                successful += 1
            except Exception:
                failed += 1

        return {"successful": successful, "failed": failed}

    def _bulk_approve_payslips(self, payslip_ids, user):
        successful = 0

        for payslip_id in payslip_ids:
            try:
                payslip = Payslip.objects.get(id=payslip_id)
                if payslip.status == "CALCULATED":
                    payslip.approve(user)
                    successful += 1
            except Exception:
                pass

        return {"successful": successful}

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_authenticated

    def has_delete_permission(self, request, obj=None):
        return False


class PayrollSystemStatusAdmin(admin.ModelAdmin):
    change_list_template = "admin/payroll/system_status.html"

    def changelist_view(self, request, extra_context=None):
        try:
            system_health = self._get_system_health()
            service_status = self._get_service_status()

            active_periods = PayrollPeriod.objects.filter(
                status__in=["DRAFT", "PROCESSING"]
            ).count()
            pending_advances = SalaryAdvance.objects.filter(status="PENDING").count()
            failed_transfers = PayrollBankTransfer.objects.filter(
                status="FAILED"
            ).count()

            extra_context = extra_context or {}
            extra_context.update(
                {
                    "system_health": system_health,
                    "service_status": service_status,
                    "active_periods": active_periods,
                    "pending_advances": pending_advances,
                    "failed_transfers": failed_transfers,
                    "title": "System Status",
                    "has_add_permission": False,
                    "has_change_permission": False,
                    "has_delete_permission": False,
                }
            )
        except Exception as e:
            messages.error(request, f"Error loading system status: {str(e)}")
            extra_context = {"error": str(e)}

        return super().changelist_view(request, extra_context=extra_context)

    def _get_system_health(self):
        try:
            total_payslips = Payslip.objects.count()
            error_payslips = Payslip.objects.filter(status="CANCELLED").count()
            error_rate = (
                (error_payslips / total_payslips * 100) if total_payslips > 0 else 0
            )

            return {
                "status": (
                    "healthy"
                    if error_rate < 5
                    else "warning" if error_rate < 10 else "critical"
                ),
                "error_rate": round(error_rate, 2),
                "total_records": total_payslips,
                "error_records": error_payslips,
            }
        except Exception:
            return {"status": "unknown", "error_rate": 0}

    def _get_service_status(self):
        services = []

        try:
            PayrollPeriod.objects.count()
            services.append({"name": "Payroll Periods", "status": "operational"})
        except Exception:
            services.append({"name": "Payroll Periods", "status": "error"})

        try:
            Payslip.objects.count()
            services.append({"name": "Payslips", "status": "operational"})
        except Exception:
            services.append({"name": "Payslips", "status": "error"})

        try:
            SalaryAdvance.objects.count()
            services.append({"name": "Salary Advances", "status": "operational"})
        except Exception:
            services.append({"name": "Salary Advances", "status": "error"})

        return services

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return request.user.is_authenticated

    def has_delete_permission(self, request, obj=None):
        return False


class PayrollAdminSite(admin.AdminSite):
    site_header = "Payroll Management System"
    site_title = "Payroll Admin"
    index_title = "Payroll Administration"
    site_url = "/payroll/"

    def index(self, request, extra_context=None):
        extra_context = extra_context or {}

        try:
            extra_context.update(
                {
                    "quick_stats": self.get_quick_stats(),
                    "system_alerts": self.get_system_alerts(),
                }
            )
        except Exception as e:
            extra_context["error"] = str(e)

        return super().index(request, extra_context)

    def get_quick_stats(self):
        try:
            current_year = timezone.now().year
            current_month = timezone.now().month

            return {
                "active_employees": CustomUser.objects.filter(
                    is_active=True, status="ACTIVE"
                ).count(),
                "current_period_exists": PayrollPeriod.objects.filter(
                    year=current_year, month=current_month
                ).exists(),
                "pending_advances": SalaryAdvance.objects.filter(
                    status="PENDING"
                ).count(),
                "draft_payslips": Payslip.objects.filter(status="DRAFT").count(),
                "processing_periods": PayrollPeriod.objects.filter(
                    status="PROCESSING"
                ).count(),
            }
        except Exception:
            return {}

    def get_system_alerts(self):
        alerts = []

        try:
            overdue_advances = SalaryAdvance.objects.filter(
                status="ACTIVE", is_overdue=True
            ).count()
            if overdue_advances > 0:
                alerts.append(
                    {
                        "type": "warning",
                        "message": f"{overdue_advances} salary advances are overdue",
                        "action_url": "/admin/payroll/salaryadvance/?status=ACTIVE",
                    }
                )

            failed_transfers = PayrollBankTransfer.objects.filter(
                status="FAILED"
            ).count()
            if failed_transfers > 0:
                alerts.append(
                    {
                        "type": "error",
                        "message": f"{failed_transfers} bank transfers have failed",
                        "action_url": "/admin/payroll/payrollbanktransfer/?status=FAILED",
                    }
                )

            incomplete_periods = PayrollPeriod.objects.filter(
                status="PROCESSING"
            ).count()
            if incomplete_periods > 0:
                alerts.append(
                    {
                        "type": "info",
                        "message": f"{incomplete_periods} payroll periods are being processed",
                        "action_url": "/admin/payroll/payrollperiod/?status=PROCESSING",
                    }
                )
        except Exception:
            pass

        return alerts


payroll_admin_site = PayrollAdminSite(name="payroll_admin")


admin.site.site_header = "HR Management System - Payroll Module"
admin.site.site_title = "Payroll Admin"
admin.site.index_title = "Payroll Management Dashboard"


def customize_admin_interface():
    admin.site.enable_nav_sidebar = True
    original_get_app_list = admin.AdminSite.get_app_list

    def get_app_list(self, request):

        app_list = original_get_app_list(self, request)

        for app in app_list:
            if app["app_label"] == "payroll":
                app["models"].sort(
                    key=lambda x: {
                        "PayrollPeriod": 1,
                        "Payslip": 2,
                        "PayslipItem": 3,
                        "SalaryAdvance": 4,
                        "PayrollDepartmentSummary": 5,
                        "PayrollBankTransfer": 6,
                    }.get(x["object_name"], 999)
                )

        return app_list
    
    admin.AdminSite.get_app_list = get_app_list

def register_admin_actions():
    def make_active(modeladmin, request, queryset):
        if hasattr(queryset.model, "is_active"):
            updated = queryset.update(is_active=True)
            messages.success(request, f"Activated {updated} items")

    make_active.short_description = "Activate selected items"

    def make_inactive(modeladmin, request, queryset):
        if hasattr(queryset.model, "is_active"):
            updated = queryset.update(is_active=False)
            messages.success(request, f"Deactivated {updated} items")

    make_inactive.short_description = "Deactivate selected items"

    admin.site.add_action(make_active)
    admin.site.add_action(make_inactive)
    admin.site.add_action(export_to_excel)


class PayrollAdminUtils:
    @staticmethod
    def format_currency(amount):
        return f"LKR {amount:,.2f}" if amount else "LKR 0.00"

    @staticmethod
    def format_percentage(value):
        return f"{value:.1f}%" if value else "0.0%"

    @staticmethod
    def get_status_color(status):
        colors = {
            "DRAFT": "#6c757d",
            "PENDING": "#ffc107",
            "PROCESSING": "#007bff",
            "CALCULATED": "#17a2b8",
            "APPROVED": "#28a745",
            "COMPLETED": "#28a745",
            "ACTIVE": "#28a745",
            "PAID": "#6f42c1",
            "CANCELLED": "#dc3545",
            "FAILED": "#dc3545",
            "GENERATED": "#007bff",
            "SENT": "#ffc107",
        }
        return colors.get(status, "#6c757d")


def setup_payroll_admin():
    customize_admin_interface()
    register_admin_actions()


def admin_view_decorator(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("admin:login")
        return view_func(request, *args, **kwargs)

    return wrapper


class PayrollAdminMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        response = self.get_response(request)
        return response


try:
    setup_payroll_admin()
except Exception as e:
    import logging

    logger = logging.getLogger(__name__)
    logger.error(f"Error setting up payroll admin: {str(e)}")
