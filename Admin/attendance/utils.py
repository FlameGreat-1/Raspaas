from django.utils import timezone
from django.db.models import Q, Sum, Count
from django.core.exceptions import ValidationError
from accounts.models import CustomUser, SystemConfiguration
from employees.models import EmployeeProfile, Contract
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, time, timedelta
import socket
import struct
import json
import logging
import hashlib
import uuid
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import io
import base64

logger = logging.getLogger(__name__)

class TimeCalculator:
    @staticmethod
    def parse_time_string(time_str: str) -> Optional[time]:
        if not time_str or time_str.strip() == '':
            return None
        
        try:
            if ':' in time_str:
                parts = time_str.split(':')
                if len(parts) == 2:
                    hour, minute = int(parts[0]), int(parts[1])
                    return time(hour, minute, 0)
                elif len(parts) == 3:
                    hour, minute, second = int(parts[0]), int(parts[1]), int(parts[2])
                    return time(hour, minute, second)
            return None
        except (ValueError, TypeError):
            return None

    @staticmethod
    def time_to_decimal_hours(time_obj: time) -> Decimal:
        if not time_obj:
            return Decimal('0.00')
        
        total_seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second
        hours = Decimal(total_seconds) / Decimal('3600')
        return hours.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    @staticmethod
    def decimal_hours_to_time(decimal_hours: Decimal) -> time:
        if not decimal_hours or decimal_hours <= 0:
            return time(0, 0, 0)
        
        total_seconds = int(decimal_hours * 3600)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        
        hours = min(hours, 23)
        return time(hours, minutes, seconds)

    @staticmethod
    def duration_to_decimal_hours(duration: timedelta) -> Decimal:
        if not duration:
            return Decimal('0.00')
        
        total_seconds = duration.total_seconds()
        hours = Decimal(total_seconds) / Decimal('3600')
        return hours.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    @staticmethod
    def decimal_hours_to_duration(decimal_hours: Decimal) -> timedelta:
        if not decimal_hours or decimal_hours <= 0:
            return timedelta(0)
        
        total_seconds = float(decimal_hours * 3600)
        return timedelta(seconds=total_seconds)

    @staticmethod
    def calculate_time_difference(start_time: time, end_time: time) -> timedelta:
        if not start_time or not end_time:
            return timedelta(0)
        
        start_datetime = datetime.combine(date.today(), start_time)
        end_datetime = datetime.combine(date.today(), end_time)
        
        if end_datetime < start_datetime:
            end_datetime += timedelta(days=1)
        
        return end_datetime - start_datetime

    @staticmethod
    def format_duration_to_excel_time(duration: timedelta) -> str:
        if not duration:
            return "00:00:00"
        
        total_seconds = int(duration.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    @staticmethod
    def calculate_multiple_periods(time_pairs: List[Tuple[time, time]]) -> Dict[str, timedelta]:
        total_time = timedelta(0)
        break_time = timedelta(0)
        
        valid_pairs = [(start, end) for start, end in time_pairs if start and end]
        
        for i, (start, end) in enumerate(valid_pairs):
            period_duration = TimeCalculator.calculate_time_difference(start, end)
            total_time += period_duration
            
            if i < len(valid_pairs) - 1:
                current_out = end
                next_in = valid_pairs[i + 1][0]
                if current_out and next_in:
                    break_duration = TimeCalculator.calculate_time_difference(current_out, next_in)
                    break_time += break_duration
        
        work_time = total_time - break_time if total_time > break_time else total_time
        
        return {
            'total_time': total_time,
            'break_time': break_time,
            'work_time': work_time
        }

class EmployeeDataManager:
    @staticmethod
    def get_active_employees() -> List[CustomUser]:
        return CustomUser.active.select_related(
            'department', 'role', 'employee_profile', 'manager'
        ).prefetch_related('contracts').all()

    @staticmethod
    def get_employee_by_code(employee_code: str) -> Optional[CustomUser]:
        try:
            return CustomUser.objects.select_related(
                'department', 'role', 'employee_profile'
            ).get(employee_code=employee_code, is_active=True)
        except CustomUser.DoesNotExist:
            return None

    @staticmethod
    def get_employee_profile(employee: CustomUser) -> Optional[EmployeeProfile]:
        try:
            return employee.employee_profile
        except EmployeeProfile.DoesNotExist:
            return None

    @staticmethod
    def get_employee_current_contract(employee: CustomUser) -> Optional[Contract]:
        today = timezone.now().date()
        try:
            return employee.contracts.filter(
                Q(end_date__isnull=True) | Q(end_date__gte=today),
                status='ACTIVE',
                start_date__lte=today,
            ).first()
        except Contract.DoesNotExist:
            return None

    @staticmethod
    def get_employee_work_schedule(employee: CustomUser) -> Dict[str, Any]:
        profile = EmployeeDataManager.get_employee_profile(employee)
        contract = EmployeeDataManager.get_employee_current_contract(employee)
        
        default_schedule = {
            'reporting_time': time(9, 0, 0),
            'shift_hours': Decimal('8.00'),
            'working_hours': Decimal('8.00'),
            'standard_work_time': timedelta(hours=9, minutes=15)
        }
        
        if profile:
            default_schedule.update({
                'reporting_time': profile.reporting_time,
                'shift_hours': profile.shift_hours,
            })
        
        if contract:
            default_schedule.update({
                'working_hours': contract.working_hours,
            })
        
        standard_hours = SystemConfiguration.get_float_setting('WORKING_HOURS_PER_DAY', 9.25)
        default_schedule['standard_work_time'] = timedelta(hours=int(standard_hours), 
                                                         minutes=int((standard_hours % 1) * 60))
        
        return default_schedule

    @staticmethod
    def get_employees_by_department(department_id: int) -> List[CustomUser]:
        return CustomUser.active.filter(department_id=department_id).select_related(
            'department', 'role', 'employee_profile'
        )

    @staticmethod
    def get_employee_manager_hierarchy(employee: CustomUser) -> List[CustomUser]:
        hierarchy = []
        current_manager = employee.manager
        
        while current_manager and len(hierarchy) < 10:
            hierarchy.append(current_manager)
            current_manager = current_manager.manager
        
        return hierarchy

    @staticmethod
    def validate_employee_for_attendance(employee: CustomUser) -> Tuple[bool, str]:
        if not employee.is_active:
            return False, "Employee is not active"
        
        if employee.status != 'ACTIVE':
            return False, f"Employee status is {employee.status}"
        
        profile = EmployeeDataManager.get_employee_profile(employee)
        if not profile or not profile.is_active:
            return False, "Employee profile is not active"
        
        contract = EmployeeDataManager.get_employee_current_contract(employee)
        if not contract:
            return False, "No active contract found"
        
        if contract.is_expired:
            return False, "Employee contract has expired"
        
        return True, "Valid for attendance"

class DeviceDataProcessor:
    @staticmethod
    def parse_realand_log_data(raw_data: bytes) -> Optional[Dict[str, Any]]:
        try:
            if len(raw_data) < 16:
                return None
            
            employee_id = struct.unpack('<I', raw_data[0:4])[0]
            timestamp = struct.unpack('<I', raw_data[4:8])[0]
            log_type = struct.unpack('<B', raw_data[8:9])[0]
            device_id = struct.unpack('<B', raw_data[9:10])[0]
            
            log_datetime = datetime.fromtimestamp(timestamp, tz=timezone.get_current_timezone())
            
            log_type_mapping = {
                0: 'CHECK_IN',
                1: 'CHECK_OUT',
                2: 'BREAK_START',
                3: 'BREAK_END',
                4: 'OVERTIME_IN',
                5: 'OVERTIME_OUT'
            }
            
            return {
                'employee_id': str(employee_id),
                'timestamp': log_datetime,
                'log_type': log_type_mapping.get(log_type, 'UNKNOWN'),
                'device_id': f'A-F011-{device_id}',
                'raw_data': base64.b64encode(raw_data).decode('utf-8')
            }
        except (struct.error, ValueError, OSError) as e:
            logger.error(f"Error parsing REALAND log data: {e}")
            return None

    @staticmethod
    def group_logs_by_employee_date(logs: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        grouped = {}
        
        for log in logs:
            employee_code = log.get('employee_code') or log.get('employee_id')
            log_date = log['timestamp'].date()
            key = f"{employee_code}_{log_date}"
            
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(log)
        
        for key in grouped:
            grouped[key].sort(key=lambda x: x['timestamp'])
        
        return grouped

    @staticmethod
    def create_attendance_pairs(logs: List[Dict[str, Any]]) -> List[Tuple[Optional[time], Optional[time]]]:
        pairs = []
        current_in = None
        
        for log in logs:
            log_time = log['timestamp'].time()
            log_type = log['log_type']
            
            if log_type in ['CHECK_IN', 'BREAK_END', 'OVERTIME_IN']:
                if current_in is None:
                    current_in = log_time
            elif log_type in ['CHECK_OUT', 'BREAK_START', 'OVERTIME_OUT']:
                if current_in is not None:
                    pairs.append((current_in, log_time))
                    current_in = None
                else:
                    pairs.append((None, log_time))
        
        if current_in is not None:
            pairs.append((current_in, None))
        
        while len(pairs) < 6:
            pairs.append((None, None))
        
        return pairs[:6]

    @staticmethod
    def validate_attendance_data(employee_code: str, attendance_date: date, 
                               time_pairs: List[Tuple[Optional[time], Optional[time]]]) -> Tuple[bool, List[str]]:
        errors = []
        
        employee = EmployeeDataManager.get_employee_by_code(employee_code)
        if not employee:
            errors.append(f"Employee with code {employee_code} not found")
            return False, errors
        
        is_valid, message = EmployeeDataManager.validate_employee_for_attendance(employee)
        if not is_valid:
            errors.append(message)
        
        if attendance_date > timezone.now().date():
            errors.append("Attendance date cannot be in the future")
        
        for i, (in_time, out_time) in enumerate(time_pairs):
            if in_time and out_time:
                if in_time >= out_time:
                    errors.append(f"Check-in time must be before check-out time for pair {i+1}")
            elif in_time and not out_time:
                if i < len(time_pairs) - 1:
                    errors.append(f"Incomplete attendance pair {i+1}: missing check-out time")
        
        return len(errors) == 0, errors

class AttendanceCalculator:
    @staticmethod
    def calculate_attendance_metrics(time_pairs: List[Tuple[Optional[time], Optional[time]]], 
                                   employee: CustomUser, attendance_date: date) -> Dict[str, Any]:
        
        schedule = EmployeeDataManager.get_employee_work_schedule(employee)
        standard_work_time = schedule['standard_work_time']
        
        time_calculations = TimeCalculator.calculate_multiple_periods(time_pairs)
        
        total_time = time_calculations['total_time']
        break_time = time_calculations['break_time']
        work_time = time_calculations['work_time']
        
        overtime = timedelta(0)
        if work_time > standard_work_time:
            overtime = work_time - standard_work_time
        
        undertime = timedelta(0)
        if work_time < standard_work_time:
            undertime = standard_work_time - work_time
        
        first_in = None
        last_out = None
        
        for in_time, out_time in time_pairs:
            if in_time and first_in is None:
                first_in = in_time
            if out_time:
                last_out = out_time
        
        status = AttendanceCalculator.determine_attendance_status(
            time_pairs, first_in, last_out, schedule, attendance_date
        )
        
        late_minutes = AttendanceCalculator.calculate_late_minutes(
            first_in, schedule['reporting_time']
        )
        
        early_departure_minutes = AttendanceCalculator.calculate_early_departure_minutes(
            last_out, schedule, work_time
        )
        
        return {
            'total_time': total_time,
            'break_time': break_time,
            'work_time': work_time,
            'overtime': overtime,
            'undertime': undertime,
            'first_in_time': first_in,
            'last_out_time': last_out,
            'status': status,
            'late_minutes': late_minutes,
            'early_departure_minutes': early_departure_minutes,
            'is_complete_day': AttendanceCalculator.is_complete_working_day(work_time, standard_work_time),
            'attendance_percentage': AttendanceCalculator.calculate_attendance_percentage(work_time, standard_work_time)
        }
    
    @staticmethod
    def determine_attendance_status(time_pairs: List[Tuple[Optional[time], Optional[time]]], 
                                  first_in: Optional[time], last_out: Optional[time],
                                  schedule: Dict[str, Any], attendance_date: date) -> str:
        
        if not first_in:
            return 'ABSENT'
        
        grace_period_minutes = SystemConfiguration.get_int_setting('GRACE_PERIOD_MINUTES', 15)
        half_day_threshold_hours = SystemConfiguration.get_float_setting('HALF_DAY_THRESHOLD_HOURS', 4.0)
        
        reporting_time = schedule['reporting_time']
        grace_time = (datetime.combine(date.today(), reporting_time) + 
                     timedelta(minutes=grace_period_minutes)).time()
        
        work_time_calculations = TimeCalculator.calculate_multiple_periods(time_pairs)
        work_hours = TimeCalculator.duration_to_decimal_hours(work_time_calculations['work_time'])
        
        if work_hours < Decimal(str(half_day_threshold_hours)):
            return 'HALF_DAY'
        
        if first_in > grace_time:
            return 'LATE'
        
        if not last_out:
            return 'INCOMPLETE'
        
        return 'PRESENT'
    
    @staticmethod
    def calculate_late_minutes(actual_in_time: Optional[time], expected_in_time: time) -> int:
        if not actual_in_time:
            return 0
        
        if actual_in_time <= expected_in_time:
            return 0
        
        actual_datetime = datetime.combine(date.today(), actual_in_time)
        expected_datetime = datetime.combine(date.today(), expected_in_time)
        
        late_duration = actual_datetime - expected_datetime
        return int(late_duration.total_seconds() / 60)
    
    @staticmethod
    def calculate_early_departure_minutes(actual_out_time: Optional[time], 
                                        schedule: Dict[str, Any], work_time: timedelta) -> int:
        if not actual_out_time:
            return 0
        
        standard_work_time = schedule['standard_work_time']
        reporting_time = schedule['reporting_time']
        
        expected_out_time = (datetime.combine(date.today(), reporting_time) + standard_work_time).time()
        
        if actual_out_time >= expected_out_time:
            return 0
        
        if work_time >= standard_work_time:
            return 0
        
        actual_datetime = datetime.combine(date.today(), actual_out_time)
        expected_datetime = datetime.combine(date.today(), expected_out_time)
        
        early_duration = expected_datetime - actual_datetime
        return int(early_duration.total_seconds() / 60)
    
    @staticmethod
    def is_complete_working_day(work_time: timedelta, standard_work_time: timedelta) -> bool:
        completion_threshold = SystemConfiguration.get_float_setting('COMPLETION_THRESHOLD_PERCENTAGE', 90.0)
        
        if standard_work_time.total_seconds() == 0:
            return False
        
        completion_percentage = (work_time.total_seconds() / standard_work_time.total_seconds()) * 100
        return completion_percentage >= completion_threshold
    
    @staticmethod
    def calculate_attendance_percentage(work_time: timedelta, standard_work_time: timedelta) -> Decimal:
        if standard_work_time.total_seconds() == 0:
            return Decimal('0.00')
        
        percentage = (work_time.total_seconds() / standard_work_time.total_seconds()) * 100
        return Decimal(str(min(percentage, 100.0))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

class MonthlyCalculator:
    @staticmethod
    def calculate_monthly_summary(employee: CustomUser, year: int, month: int) -> Dict[str, Any]:
        from .models import Attendance
        
        monthly_records = Attendance.objects.filter(
            employee=employee,
            date__year=year,
            date__month=month
        ).order_by('date')
        
        total_work_time = timedelta(0)
        total_break_time = timedelta(0)
        total_overtime = timedelta(0)
        total_undertime = timedelta(0)
        
        working_days = 0
        attended_days = 0
        half_days = 0
        late_days = 0
        early_days = 0
        absent_days = 0
        
        earliest_in_time = None
        latest_out_time = None
        
        for record in monthly_records:
            working_days += 1
            
            if record.status != 'ABSENT':
                attended_days += 1
                total_work_time += record.work_time
                total_break_time += record.break_time
                total_overtime += record.overtime
                total_undertime += record.undertime
                
                if record.first_in_time:
                    if earliest_in_time is None or record.first_in_time < earliest_in_time:
                        earliest_in_time = record.first_in_time
                
                if record.last_out_time:
                    if latest_out_time is None or record.last_out_time > latest_out_time:
                        latest_out_time = record.last_out_time
            
            if record.status == 'HALF_DAY':
                half_days += 1
            elif record.status == 'LATE':
                late_days += 1
            elif record.status == 'ABSENT':
                absent_days += 1
            
            if record.early_departure_minutes > 0:
                early_days += 1
        
        attendance_percentage = Decimal('0.00')
        if working_days > 0:
            attendance_percentage = (Decimal(attended_days) / Decimal(working_days) * 100).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            )
        
        return {
            'employee': employee,
            'year': year,
            'month': month,
            'total_work_time': total_work_time,
            'total_break_time': total_break_time,
            'total_overtime': total_overtime,
            'total_undertime': total_undertime,
            'working_days': working_days,
            'attended_days': attended_days,
            'half_days': half_days,
            'late_days': late_days,
            'early_days': early_days,
            'absent_days': absent_days,
            'leave_days': 0,
            'holiday_days': 0,
            'attendance_percentage': attendance_percentage,
            'earliest_in_time': earliest_in_time,
            'latest_out_time': latest_out_time,
            'average_work_hours': MonthlyCalculator.calculate_average_work_hours(total_work_time, attended_days),
            'punctuality_score': MonthlyCalculator.calculate_punctuality_score(late_days, early_days, attended_days)
        }
    
    @staticmethod
    def calculate_average_work_hours(total_work_time: timedelta, attended_days: int) -> Decimal:
        if attended_days == 0:
            return Decimal('0.00')
        
        average_seconds = total_work_time.total_seconds() / attended_days
        average_hours = Decimal(average_seconds) / Decimal('3600')
        return average_hours.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    @staticmethod
    def calculate_punctuality_score(late_days: int, early_days: int, attended_days: int) -> Decimal:
        if attended_days == 0:
            return Decimal('100.00')
        
        punctual_days = attended_days - late_days - early_days
        score = (Decimal(punctual_days) / Decimal(attended_days) * 100).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
        return max(score, Decimal('0.00'))

class ExcelProcessor:
    @staticmethod
    def create_attendance_excel(employee_data: List[Dict[str, Any]], month: int, year: int) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance_{year}_{month:02d}"
        
        headers = [
            'Division', 'ID', 'Name', 'In1', 'Out1', 'In2', 'Out2', 'In3', 'Out3',
            'In4', 'Out4', 'In5', 'Out5', 'In6', 'Out6', 'Total', 'Break', 'Work', 'Over'
        ]
        
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        for row, data in enumerate(employee_data, 2):
            ws.cell(row=row, column=1, value=data.get('division', ''))
            ws.cell(row=row, column=2, value=data.get('employee_id', ''))
            ws.cell(row=row, column=3, value=data.get('name', ''))
            
            time_pairs = data.get('time_pairs', [])
            for i, (in_time, out_time) in enumerate(time_pairs):
                in_col = 4 + (i * 2)
                out_col = 5 + (i * 2)
                
                ws.cell(row=row, column=in_col, value=in_time.strftime('%H:%M:%S') if in_time else '')
                ws.cell(row=row, column=out_col, value=out_time.strftime('%H:%M:%S') if out_time else '')
            
            ws.cell(row=row, column=16, value=TimeCalculator.format_duration_to_excel_time(data.get('total_time', timedelta(0))))
            ws.cell(row=row, column=17, value=TimeCalculator.format_duration_to_excel_time(data.get('break_time', timedelta(0))))
            ws.cell(row=row, column=18, value=TimeCalculator.format_duration_to_excel_time(data.get('work_time', timedelta(0))))
            ws.cell(row=row, column=19, value=TimeCalculator.format_duration_to_excel_time(data.get('overtime', timedelta(0))))
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

class DeviceManager:
    @staticmethod
    def connect_to_realand_device(device_ip: str, device_port: int = 4370) -> Optional[socket.socket]:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(30)
            sock.connect((device_ip, device_port))
            return sock
        except (socket.error, socket.timeout) as e:
            logger.error(f"Failed to connect to REALAND device {device_ip}:{device_port} - {e}")
            return None

    @staticmethod
    def send_command_to_device(sock: socket.socket, command: bytes) -> Optional[bytes]:
        try:
            sock.send(command)
            response = sock.recv(1024)
            return response
        except socket.error as e:
            logger.error(f"Failed to send command to device: {e}")
            return None

    @staticmethod
    def sync_employee_to_device(sock: socket.socket, employee: CustomUser) -> bool:
        try:
            employee_data = struct.pack(
                '<I50s50s',
                int(employee.employee_code.replace('EMP', '').replace('ADMIN', '999')),
                employee.get_full_name().encode('utf-8')[:50].ljust(50, b'\x00'),
                employee.employee_code.encode('utf-8')[:50].ljust(50, b'\x00')
            )
            
            command = b'\x01' + employee_data
            response = DeviceManager.send_command_to_device(sock, command)
            
            return response is not None and len(response) > 0
        except (struct.error, UnicodeEncodeError) as e:
            logger.error(f"Failed to sync employee {employee.employee_code} to device: {e}")
            return False

    @staticmethod
    def get_attendance_logs_from_device(sock: socket.socket, start_date: date, end_date: date) -> List[Dict[str, Any]]:
        try:
            start_timestamp = int(datetime.combine(start_date, time.min).timestamp())
            end_timestamp = int(datetime.combine(end_date, time.max).timestamp())
            
            command = struct.pack('<BII', 0x02, start_timestamp, end_timestamp)
            response = DeviceManager.send_command_to_device(sock, command)
            
            if not response or len(response) < 4:
                return []
            
            record_count = struct.unpack('<I', response[:4])[0]
            logs = []
            
            for i in range(record_count):
                offset = 4 + (i * 16)
                if offset + 16 <= len(response):
                    log_data = response[offset:offset + 16]
                    parsed_log = DeviceDataProcessor.parse_realand_log_data(log_data)
                    if parsed_log:
                        logs.append(parsed_log)
            
            return logs
        except (struct.error, ValueError) as e:
            logger.error(f"Failed to get attendance logs from device: {e}")
            return []

    @staticmethod
    def test_device_connection(device_ip: str, device_port: int = 4370) -> Tuple[bool, str]:
        sock = DeviceManager.connect_to_realand_device(device_ip, device_port)
        if not sock:
            return False, f"Cannot connect to device at {device_ip}:{device_port}"
        
        try:
            ping_command = b'\x00\x00\x00\x00'
            response = DeviceManager.send_command_to_device(sock, ping_command)
            
            if response:
                return True, "Device connection successful"
            else:
                return False, "Device did not respond to ping"
        finally:
            sock.close()

class ValidationHelper:
    @staticmethod
    def validate_time_format(time_str: str) -> Tuple[bool, str]:
        if not time_str or time_str.strip() == '':
            return True, "Empty time is valid"
        
        try:
            time_obj = TimeCalculator.parse_time_string(time_str)
            if time_obj is None:
                return False, "Invalid time format. Use HH:MM or HH:MM:SS"
            return True, "Valid time format"
        except Exception as e:
            return False, f"Time validation error: {str(e)}"

    @staticmethod
    def validate_date_range(start_date: date, end_date: date) -> Tuple[bool, str]:
        if start_date > end_date:
            return False, "Start date cannot be after end date"
        
        if end_date > timezone.now().date():
            return False, "End date cannot be in the future"
        
        max_range_days = SystemConfiguration.get_int_setting('MAX_DATE_RANGE_DAYS', 365)
        if (end_date - start_date).days > max_range_days:
            return False, f"Date range cannot exceed {max_range_days} days"
        
        return True, "Valid date range"

    @staticmethod
    def validate_excel_file(file_content: bytes) -> Tuple[bool, str, Optional[pd.DataFrame]]:
        try:
            df = pd.read_excel(io.BytesIO(file_content))
            
            required_columns = ['Division', 'ID', 'Name']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}", None
            
            if df.empty:
                return False, "Excel file is empty", None
            
            return True, "Valid Excel file", df
        except Exception as e:
            return False, f"Excel validation error: {str(e)}", None

    @staticmethod
    def sanitize_employee_code(employee_code: str) -> str:
        if not employee_code:
            return ""
        
        sanitized = ''.join(c for c in employee_code.upper() if c.isalnum())
        return sanitized[:20]

    @staticmethod
    def validate_attendance_consistency(time_pairs: List[Tuple[Optional[time], Optional[time]]]) -> Tuple[bool, List[str]]:
        errors = []
        last_out_time = None
        
        for i, (in_time, out_time) in enumerate(time_pairs):
            if not in_time and not out_time:
                continue
            
            if in_time and not out_time and i < len(time_pairs) - 1:
                next_pair_has_data = any(time_pairs[j][0] or time_pairs[j][1] for j in range(i + 1, len(time_pairs)))
                if next_pair_has_data:
                    errors.append(f"Incomplete pair {i + 1}: missing check-out time")
            
            if in_time and out_time and in_time >= out_time:
                errors.append(f"Check-in time must be before check-out time for pair {i + 1}")
            
            if last_out_time and in_time and in_time < last_out_time:
                errors.append(f"Check-in time for pair {i + 1} cannot be before previous check-out time")
            
            if out_time:
                last_out_time = out_time
        
        return len(errors) == 0, errors

class CacheManager:
    @staticmethod
    def get_cache_key(prefix: str, *args) -> str:
        key_parts = [prefix] + [str(arg) for arg in args]
        key_string = '_'.join(key_parts)
        return hashlib.md5(key_string.encode()).hexdigest()

    @staticmethod
    def cache_employee_schedule(employee_id: int, schedule_data: Dict[str, Any], timeout: int = 3600):
        from django.core.cache import cache
        cache_key = CacheManager.get_cache_key('employee_schedule', employee_id)
        cache.set(cache_key, schedule_data, timeout)

    @staticmethod
    def get_cached_employee_schedule(employee_id: int) -> Optional[Dict[str, Any]]:
        from django.core.cache import cache
        cache_key = CacheManager.get_cache_key('employee_schedule', employee_id)
        return cache.get(cache_key)

    @staticmethod
    def invalidate_employee_cache(employee_id: int):
        from django.core.cache import cache
        cache_key = CacheManager.get_cache_key('employee_schedule', employee_id)
        cache.delete(cache_key)

    @staticmethod
    def cache_monthly_summary(employee_id: int, year: int, month: int, summary_data: Dict[str, Any]):
        from django.core.cache import cache
        cache_key = CacheManager.get_cache_key('monthly_summary', employee_id, year, month)
        cache.set(cache_key, summary_data, 86400)

    @staticmethod
    def get_cached_monthly_summary(employee_id: int, year: int, month: int) -> Optional[Dict[str, Any]]:
        from django.core.cache import cache
        cache_key = CacheManager.get_cache_key('monthly_summary', employee_id, year, month)
        return cache.get(cache_key)

class AuditHelper:
    @staticmethod
    def log_attendance_change(user: CustomUser, action: str, employee: CustomUser, 
                            attendance_date: date, changes: Dict[str, Any], request=None):
        from accounts.models import AuditLog
        
        ip_address = '127.0.0.1'
        user_agent = 'System'
        
        if request:
            ip_address = request.META.get('REMOTE_ADDR', '127.0.0.1')
            user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')
        
        description = f"Attendance {action} for {employee.get_full_name()} on {attendance_date}"
        
        additional_data = {
            'employee_code': employee.employee_code,
            'attendance_date': attendance_date.isoformat(),
            'changes': changes,
            'module': 'attendance'
        }
        
        AuditLog.log_action(
            user=user,
            action=f'ATTENDANCE_{action.upper()}',
            description=description,
            ip_address=ip_address,
            user_agent=user_agent,
            additional_data=additional_data
        )

    @staticmethod
    def log_device_sync(user: CustomUser, device_id: str, sync_result: Dict[str, Any], request=None):
        from accounts.models import AuditLog
        
        ip_address = '127.0.0.1'
        user_agent = 'System'
        
        if request:
            ip_address = request.META.get('REMOTE_ADDR', '127.0.0.1')
            user_agent = request.META.get('HTTP_USER_AGENT', 'Unknown')
        
        description = f"Device sync completed for {device_id}"
        
        additional_data = {
            'device_id': device_id,
            'sync_result': sync_result,
            'module': 'attendance'
        }
        
        AuditLog.log_action(
            user=user,
            action='DEVICE_SYNC',
            description=description,
            ip_address=ip_address,
            user_agent=user_agent,
            additional_data=additional_data
        )

class ReportGenerator:
    @staticmethod
    def generate_attendance_report_data(employees: List[CustomUser], start_date: date, 
                                      end_date: date) -> List[Dict[str, Any]]:
        from .models import Attendance
        
        report_data = []
        
        for employee in employees:
            attendance_records = Attendance.objects.filter(
                employee=employee,
                date__range=[start_date, end_date]
            ).order_by('date')
            
            total_days = (end_date - start_date).days + 1
            present_days = attendance_records.filter(status__in=['PRESENT', 'LATE']).count()
            absent_days = attendance_records.filter(status='ABSENT').count()
            half_days = attendance_records.filter(status='HALF_DAY').count()
            late_days = attendance_records.filter(status='LATE').count()
            
            total_work_time = sum(
                (record.work_time for record in attendance_records),
                timedelta(0)
            )
            
            total_overtime = sum(
                (record.overtime for record in attendance_records),
                timedelta(0)
            )
            
            attendance_percentage = (present_days / total_days * 100) if total_days > 0 else 0
            
            report_data.append({
                'employee': employee,
                'employee_code': employee.employee_code,
                'employee_name': employee.get_full_name(),
                'department': employee.department.name if employee.department else 'N/A',
                'total_days': total_days,
                'present_days': present_days,
                'absent_days': absent_days,
                'half_days': half_days,
                'late_days': late_days,
                'total_work_time': total_work_time,
                'total_overtime': total_overtime,
                'attendance_percentage': round(attendance_percentage, 2),
                'average_daily_hours': TimeCalculator.duration_to_decimal_hours(total_work_time) / max(present_days, 1)
            })
        
        return report_data

    @staticmethod
    def format_duration_for_display(duration: timedelta) -> str:
        if not duration:
            return "00:00:00"
        
        total_seconds = int(abs(duration.total_seconds()))
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        
        sign = "-" if duration.total_seconds() < 0 else ""
        return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"

def get_system_timezone():
    return timezone.get_current_timezone()

def get_current_date():
    return timezone.now().date()

def get_current_datetime():
    return timezone.now()

def generate_unique_id():
    return str(uuid.uuid4())

def safe_decimal_conversion(value: Any, default: Decimal = Decimal('0.00')) -> Decimal:
    try:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if isinstance(value, str) and value.strip():
            return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return default
    except (ValueError, TypeError, InvalidOperation):
        return default

def safe_time_conversion(value: Any) -> Optional[time]:
    if isinstance(value, time):
        return value
    if isinstance(value, str):
        return TimeCalculator.parse_time_string(value)
    return None

def safe_date_conversion(value: Any) -> Optional[date]:
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            try:
                return datetime.strptime(value, '%d/%m/%Y').date()
            except ValueError:
                return None
    return None
